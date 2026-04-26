import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, "/Users/ericye/Desktop/AI自动化/retail_ops_system/backend")

from app.api.routes import _build_bale_sales_order_workbook, _build_bale_sales_pricing_workbook
from app.core.config import settings
from app.core.state import InMemoryState


class BaleSalesStateTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.original_state_file = settings.state_file
        settings.state_file = Path(self.temp_dir.name) / "runtime_state.json"
        self.state = InMemoryState()

    def tearDown(self):
        settings.state_file = self.original_state_file
        self.temp_dir.cleanup()

    def _create_sales_pool_bales(self):
        source_pool_token = "CN-SRC-BSALE-01"
        source_bale_token = f"{source_pool_token}-001"
        self.state.create_or_update_china_source_record(
            {
                "source_pool_token": source_pool_token,
                "container_type": "40HQ",
                "customs_notice_no": "BSALE240426",
                "lines": [
                    {
                        "source_bale_token": source_bale_token,
                        "supplier_name": "Youxun Demo",
                        "category_main": "dress",
                        "category_sub": "2 pieces",
                        "package_count": 2,
                        "unit_weight_kg": 40,
                        "unit_cost_amount": 100,
                        "unit_cost_currency": "KES",
                    }
                ],
            },
            created_by="warehouse_clerk_1",
        )
        self.state.update_china_source_cost(
            source_pool_token,
            {
                "cost_entries": {
                    "head_transport": {
                        "currency": "KES",
                        "amount": 40,
                        "payment_method": "bank",
                        "payer": "finance_1",
                        "payment_reference": "HEAD-01",
                        "documents": [],
                    },
                    "customs_clearance": {
                        "currency": "KES",
                        "amount": 20,
                        "payment_method": "cash",
                        "payer": "finance_1",
                        "payment_reference": "KRA-01",
                        "documents": [],
                    },
                    "tail_transport": {
                        "currency": "KES",
                        "amount": 20,
                        "payment_method": "bank",
                        "payer": "finance_1",
                        "payment_reference": "TAIL-01",
                        "documents": [],
                    },
                }
            },
            updated_by="warehouse_supervisor_1",
        )

        shipment = self.state.create_inbound_shipment(
            {
                "shipment_type": "sea",
                "customs_notice_no": "BSALE240426",
                "unload_date": "2026-04-26",
                "coc_goods_manifest": "bale sales source",
                "note": "",
                "coc_documents": [],
            }
        )
        self.state.create_parcel_batch(
            {
                "intake_type": "sea_freight",
                "inbound_shipment_no": shipment["shipment_no"],
                "source_bale_token": source_bale_token,
                "supplier_name": "Youxun Demo",
                "cargo_type": "summer apparel",
                "category_main": "dress",
                "category_sub": "2 pieces",
                "package_count": 2,
                "total_weight": 80,
                "received_by": "warehouse_clerk_1",
                "note": "for bale sales test",
            }
        )
        self.state.confirm_inbound_shipment_intake(
            shipment["shipment_no"],
            {
                "declared_total_packages": 2,
                "confirmed_by": "warehouse_supervisor_1",
                "note": "ready for sales pool",
            },
        )
        bales = self.state.generate_bale_barcodes(shipment["shipment_no"], "warehouse_supervisor_1")
        routed = []
        for bale in bales:
            routed.append(
                self.state.route_raw_bale_to_bale_sales_pool(
                    bale["bale_barcode"],
                    {"updated_by": "warehouse_supervisor_1", "note": "send to sales"},
                )
            )
        return shipment, routed

    def test_list_candidates_and_pricing_update_allocate_cost_and_margin(self):
        shipment, routed = self._create_sales_pool_bales()

        candidates = self.state.list_bale_sales_candidates(shipment_no=shipment["shipment_no"])
        self.assertEqual(len(candidates), 2)
        first = candidates[0]
        self.assertEqual(first["source_type"], "raw_direct_sale")
        self.assertEqual(first["shipment_no"], shipment["shipment_no"])
        self.assertAlmostEqual(first["source_cost_kes"], 140.0, places=2)
        self.assertAlmostEqual(first["editable_cost_kes"], 140.0, places=2)
        self.assertEqual(first["status"], "available")

        updated = self.state.update_bale_sales_candidate_pricing(
            first["entry_id"],
            {
                "updated_by": "admin_1",
                "editable_cost_kes": 150,
                "downstream_cost_kes": 15,
                "margin_rate": 0.2,
                "note": "add local handling",
            },
        )
        self.assertAlmostEqual(updated["editable_cost_kes"], 150.0, places=2)
        self.assertAlmostEqual(updated["downstream_cost_kes"], 15.0, places=2)
        self.assertAlmostEqual(updated["total_cost_kes"], 165.0, places=2)
        self.assertAlmostEqual(updated["margin_rate"], 0.2, places=4)
        self.assertAlmostEqual(updated["target_sale_price_kes"], 198.0, places=2)

        workbook_bytes = _build_bale_sales_pricing_workbook([updated])
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.title, "待售包裹成本表")
        self.assertEqual(sheet["A1"].value, "待售 bale 成本与毛利编辑表")
        self.assertEqual(sheet["A4"].value, "entry_id")
        self.assertEqual(sheet["J5"].value, 150)
        self.assertEqual(sheet["L5"].value, 165)
        self.assertEqual(sheet["N5"].value, 198)

    def test_complete_outbound_marks_bale_sold_and_generates_sales_workbook(self):
        shipment, _ = self._create_sales_pool_bales()
        candidates = self.state.list_bale_sales_candidates(shipment_no=shipment["shipment_no"])
        first = self.state.update_bale_sales_candidate_pricing(
            candidates[0]["entry_id"],
            {
                "updated_by": "admin_1",
                "editable_cost_kes": 150,
                "downstream_cost_kes": 15,
                "target_sale_price_kes": 210,
                "note": "approved for sale",
            },
        )

        order = self.state.create_bale_sales_order(
            {
                "created_by": "admin_1",
                "sold_by": "Austin",
                "customer_name": "Nairobi Textile Buyer",
                "customer_contact": "+254700000111",
                "payment_method": "bank_transfer",
                "note": "same-day pickup",
                "items": [
                    {
                        "entry_id": first["entry_id"],
                        "sale_price_kes": 215,
                    }
                ],
            }
        )

        self.assertEqual(order["status"], "completed")
        self.assertEqual(order["sold_by"], "Austin")
        self.assertEqual(order["customer_name"], "Nairobi Textile Buyer")
        self.assertEqual(order["customer_contact"], "+254700000111")
        self.assertEqual(order["payment_method"], "bank_transfer")
        self.assertAlmostEqual(order["total_cost_kes"], 165.0, places=2)
        self.assertAlmostEqual(order["total_amount_kes"], 215.0, places=2)
        self.assertAlmostEqual(order["total_profit_kes"], 50.0, places=2)
        self.assertEqual(order["items"][0]["bale_barcode"], first["bale_barcode"])

        refreshed = self.state.list_bale_sales_candidates(shipment_no=shipment["shipment_no"])
        sold_candidate = next(row for row in refreshed if row["entry_id"] == first["entry_id"])
        self.assertEqual(sold_candidate["status"], "sold")
        self.assertEqual(sold_candidate["outbound_order_no"], order["order_no"])
        self.assertEqual(sold_candidate["is_available"], False)

        raw_bale = self.state.list_raw_bales(
            shipment_no=shipment["shipment_no"],
            destination_judgement="bale_sales_pool",
        )[0]
        self.assertEqual(raw_bale["status"], "sold_via_bale_sales")

        workbook_bytes = _build_bale_sales_order_workbook(order)
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.title, "Bales销售单")
        self.assertEqual(sheet["A1"].value, "Bales 销售出库单")
        self.assertEqual(sheet["B2"].value, order["order_no"])
        self.assertEqual(sheet["B4"].value, "Nairobi Textile Buyer")
        self.assertEqual(sheet["A8"].value, "bale_barcode")
        self.assertEqual(sheet["E9"].value, 215)


if __name__ == "__main__":
    unittest.main()
