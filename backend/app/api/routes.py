import csv
import hashlib
import json
import math
import os
import re
import subprocess
from datetime import datetime, timezone
from io import BytesIO, StringIO
from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile
from textwrap import wrap
from typing import Any, Optional

from fastapi import APIRouter, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import createBarcodeDrawing
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from app.core.config import settings
from app.core.seed_data import (
    BARCODE_SETTINGS,
    ROLE_DEFINITIONS,
    STORE_RACK_TEMPLATE,
)
from app.core.state import state
from app.schemas.adjustments import InventoryAdjustmentResponse
from app.schemas.audit import AuditEventResponse
from app.schemas.auth import LoginRequest, LoginResponse, SessionUserResponse
from app.schemas.bale_sales import (
    BaleSalesCandidatePricingUpdateRequest,
    BaleSalesCandidateResponse,
    BaleSalesOrderCreate,
    BaleSalesOrderResponse,
)
from app.schemas.barcodes import BarcodeResolveResponse
from app.schemas.cargo_types import CargoTypeCreate, CargoTypeResponse
from app.schemas.common import MessageResponse, SummaryCard
from app.schemas.identity import ItemIdentityLedgerResponse
from app.schemas.integrations import (
    MpesaCallbackResponse,
    MpesaCollectionImportRequest,
    MpesaCollectionResponse,
    MpesaCustomerInsightResponse,
    OfflineSaleSyncBatchRequest,
    OfflineSaleSyncBatchResponse,
)
from app.schemas.movements import InventoryMovementResponse
from app.schemas.payments import PaymentAnomalyResolveRequest, PaymentAnomalyResponse
from app.schemas.printing import (
    BaleLabelPrintJobCreate,
    BaleLabelPrintJobResponse,
    BaleDirectPrintBatchRequest,
    BaleDirectPrintRequest,
    BaleBarcodePrintRequest,
    BaleBarcodePrintResponse,
    DocumentPrintJobCreate,
    ItemBarcodeTokenPrintJobCreate,
    LabelCandidateBatchPrintRequest,
    LabelCandidatePrintRequest,
    LabelPrintJobCreate,
    PrintStationClaimRequest,
    PrintStationCompleteRequest,
    PrintStationFailRequest,
    PrintJobFailureRequest,
    PrintJobResponse,
    StorePrepBalePrintJobCreate,
    SystemPrinterResponse,
    TransferDispatchBundleRequest,
    TransferDispatchBundleResponse,
)
from app.schemas.pos import (
    CashierHandoverLogResponse,
    CashierHandoverRequest,
    CashierHandoverReviewRequest,
    CashierShiftCloseRequest,
    CashierShiftOpenRequest,
    StoreClosingChecklistResponse,
    CashierShiftSummary,
    PosReportResponse,
)
from app.schemas.pricing import PriceRuleCreate, PriceRuleResponse
from app.schemas.products import (
    ProductBarcodeAssignRequest,
    ProductBulkCreateRequest,
    ProductCreate,
    ProductImportPreviewResponse,
    ProductImportRow,
    ProductResponse,
)
from app.schemas.receiving import (
    PlacementSuggestionResponse,
    ReceivingBatchAddRequest,
    ReceivingSessionFinalizeRequest,
    ReceivingSessionResponse,
    ReceivingSessionStartRequest,
    StoreTokenPlacementSuggestionResponse,
    StoreTokenReceivingBatchAddRequest,
    StoreTokenReceivingBatchResponse,
    StoreTokenReceivingSessionResponse,
    StoreTokenReceivingSessionStartRequest,
)
from app.schemas.refunds import (
    SaleRefundRequestCreate,
    SaleRefundRequestResponse,
    SaleRefundReviewRequest,
)
from app.schemas.sales import (
    RecentStoreSalesSimulationRequest,
    RecentStoreSalesSimulationResponse,
    SaleCreate,
    SaleResponse,
    StoreRetailSeedRequest,
    StoreRetailSeedResponse,
)
from app.schemas.sorting import (
    ApparelPieceWeightCreate,
    ApparelDefaultCostCreate,
    ApparelDefaultCostResponse,
    ApparelSortingRackCreate,
    ApparelSortingRackResponse,
    ApparelPieceWeightResponse,
    BaleBarcodeResponse,
    ChinaSourceCostUpdateRequest,
    ChinaSourceRecordCreate,
    ChinaSourceRecordResponse,
    ChinaSourceImportPreviewResponse,
    ChinaSourceImportRow,
    InboundShipmentCreate,
    InboundShipmentIntakeConfirmRequest,
    InboundShipmentResponse,
    ItemBarcodeTokenResponse,
    ItemBarcodeTokenStoreEditRequest,
    ParcelBatchCreate,
    ParcelBatchResponse,
    RawBaleRouteRequest,
    RawBaleStockResponse,
    StorePrepBaleResponse,
    StorePrepBaleTaskCompleteRequest,
    StorePrepBaleTaskCreate,
    StorePrepBaleTaskResponse,
    SortingStockRackUpdateRequest,
    SortingStockRackUpdateResponse,
    SortingStockResponse,
    StoreReplenishmentDemoResponse,
    WarehouseInventorySummaryResponse,
    WarehouseMainflowDemoResponse,
    StoreDispatchBaleAcceptRequest,
    StoreDispatchBaleAssignRequest,
    StoreDispatchBaleResponse,
    SortingTaskCreate,
    SortingTaskResponse,
    SortingTaskResultSubmit,
)
from app.schemas.returns import (
    ReturnCandidateResponse,
    ReturnOrderCreate,
    ReturnOrderDispatchRequest,
    ReturnOrderReceiveRequest,
    ReturnOrderResponse,
    ReturnSelectionCreate,
)
from app.schemas.store_racks import (
    StoreRackAssignmentRequest,
    StoreRackAssignmentResponse,
    StoreRackInitializationResponse,
    StoreRackLocationResponse,
    StoreRackTemplateResponse,
)
from app.schemas.stores import (
    BarcodeSettingsResponse,
    LabelTemplateResponse,
    LabelTemplateSaveRequest,
    StoreCreate,
    StoreSiteRecommendationRequest,
    StoreSiteRecommendationResponse,
    StoreOperatingSummaryResponse,
    StoreResponse,
)
from app.schemas.suppliers import SupplierCreate, SupplierResponse
from app.schemas.transfers import (
    DiscrepancyApprovalRequest,
    PickingWaveCreate,
    PickingWaveResponse,
    RecommendationTransferCreateRequest,
    StoreDeliveryExecutionOrderCreateRequest,
    StoreDeliveryExecutionOrderResponse,
    TransferApprovalRequest,
    TransferOrderCreate,
    TransferOrderResponse,
    TransferRecommendationRequest,
    TransferRecommendationResponse,
    TransferReceiveRequest,
    TransferShipRequest,
)
from app.schemas.users import RoleResponse, UserCreate, UserResponse, UserUpdate
from app.schemas.voids import SaleVoidRequestCreate, SaleVoidRequestResponse, SaleVoidReviewRequest
from app.schemas.warehouse import (
    GoodsReceiptCreate,
    GoodsReceiptResponse,
    StoreStockLookupResponse,
    StoreStockResponse,
    WarehouseStockResponse,
)

router = APIRouter()


@router.get("/barcode/resolve/{barcode}", response_model=BarcodeResolveResponse, tags=["barcode"])
def resolve_barcode(
    barcode: str,
    context: str = Query(default=""),
    authorization: Optional[str] = Header(default=None),
) -> BarcodeResolveResponse:
    _require_current_user(authorization=authorization)
    return BarcodeResolveResponse(**state.resolve_barcode(barcode, context=context))
_TSPL_EOL = b"\r\n"

_CANDIDATE_BITMAP_FONT_PATHS = [
    "/System/Library/Fonts/SFNSMono.ttf",
    "/System/Library/Fonts/Menlo.ttc",
    "/System/Library/Fonts/HelveticaNeue.ttc",
]


def _list_system_printers() -> list[dict[str, Any]]:
    command_env = {
        **dict(os.environ),
        "LC_ALL": "C",
        "LANG": "C",
    }
    try:
        devices = subprocess.run(
            ["/usr/bin/lpstat", "-v"],
            capture_output=True,
            text=True,
            check=False,
            env=command_env,
        )
        default = subprocess.run(
            ["/usr/bin/lpstat", "-d"],
            capture_output=True,
            text=True,
            check=False,
            env=command_env,
        )
    except Exception:
        return []

    default_name = ""
    for line in (default.stdout or "").splitlines():
        line = line.strip()
        if "system default destination:" in line:
            default_name = line.split("system default destination:", 1)[1].strip()
            break
        if line.startswith("系统默认目的位置："):
            default_name = line.split("系统默认目的位置：", 1)[1].strip()
            break

    status_map: dict[str, dict[str, Any]] = {}
    page_sizes_map: dict[str, list[str]] = {}

    def _normalize_lpstat_status(line: str) -> tuple[str, bool]:
        normalized = line.strip()
        if not normalized:
            return "", False
        if "idle" in normalized.lower() or "闲置" in normalized:
            return "闲置 / 就绪", True
        if "printing" in normalized.lower() or "正在打印" in normalized:
            return "正在打印", True
        if "disabled" in normalized.lower() or "已禁用" in normalized:
            return "已禁用", False
        if "paused" in normalized.lower() or "暂停" in normalized:
            return "已暂停", False
        return normalized, False

    raw_devices: list[tuple[str, str, bool]] = []
    for raw in (devices.stdout or "").splitlines():
        line = raw.strip()
        name = ""
        uri = ""
        if line.startswith("device for "):
            name_part, _, uri = line.partition(":")
            name = name_part.replace("device for ", "").strip()
        elif line.startswith("用于") and "的设备：" in line:
            name_part, _, uri = line.partition("的设备：")
            name = name_part.replace("用于", "").strip()
        else:
            continue
        if not name:
            continue
        raw_devices.append((name, uri.strip(), name == default_name))

    for name, _, _ in raw_devices:
        try:
            status_result = subprocess.run(
                ["/usr/bin/lpstat", "-p", name, "-l"],
                capture_output=True,
                text=True,
                check=False,
                env=command_env,
            )
            status_line = next(
                (line.strip() for line in (status_result.stdout or "").splitlines() if line.strip()),
                "",
            )
            status_text, is_ready = _normalize_lpstat_status(status_line)
            status_map[name] = {
                "status_text": status_text or "未知",
                "is_ready": is_ready,
            }
        except Exception:
            status_map[name] = {
                "status_text": "未知",
                "is_ready": False,
            }

        try:
            options_result = subprocess.run(
                ["/usr/bin/lpoptions", "-p", name, "-l"],
                capture_output=True,
                text=True,
                check=False,
                env=command_env,
            )
            page_line = next(
                (
                    line.strip()
                    for line in (options_result.stdout or "").splitlines()
                    if line.strip().startswith("PageSize/")
                ),
                "",
            )
            if page_line:
                _, _, values = page_line.partition(":")
                page_sizes_map[name] = [
                    token.replace("*", "").strip()
                    for token in values.split()
                    if token.strip()
                ]
            else:
                page_sizes_map[name] = []
        except Exception:
            page_sizes_map[name] = []

    printers: list[dict[str, Any]] = []
    for name, uri, is_default in raw_devices:
        status_info = status_map.get(name, {})
        printers.append(
            {
                "name": name,
                "device_uri": uri,
                "is_default": is_default,
                "status_text": status_info.get("status_text", "未知"),
                "is_ready": bool(status_info.get("is_ready", False)),
                "supported_page_sizes": page_sizes_map.get(name, []),
            }
        )
    return printers


def _normalize_printer_name(value: str) -> str:
    return re.sub(r"[\s_-]+", "", str(value or "").strip().lower())


def _find_system_printer(printers: list[dict[str, Any]], selected_printer: str) -> Optional[dict[str, Any]]:
    requested_name = str(selected_printer or "").strip()
    if not requested_name:
        return None

    direct_map = {str(row.get("name") or "").strip(): row for row in printers}
    if requested_name in direct_map:
        return direct_map[requested_name]

    normalized_requested = _normalize_printer_name(requested_name)
    for row in printers:
        printer_name = str(row.get("name") or "").strip()
        if printer_name and _normalize_printer_name(printer_name) == normalized_requested:
            return row
    return None


def _resolve_printer_destination(printer_row: Optional[dict[str, Any]], selected_printer: str) -> str:
    actual_name = str((printer_row or {}).get("name") or "").strip()
    if actual_name:
        return actual_name
    return str(selected_printer or "").strip()


def _resolve_label_dimensions(payload: dict[str, Any], fallback_label_size: str = "60x40") -> tuple[float, float]:
    width_value = payload.get("width_mm")
    height_value = payload.get("height_mm")
    if width_value and height_value:
        try:
            return max(float(width_value), 20.0), max(float(height_value), 20.0)
        except (TypeError, ValueError):
            pass
    label_size = str(payload.get("label_size") or fallback_label_size).strip().lower()
    match = re.match(r"^\s*(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\s*$", label_size)
    if match:
        return max(float(match.group(1)), 20.0), max(float(match.group(2)), 20.0)
    return 60.0, 40.0


def _merge_print_payload_with_template(job: dict[str, Any], template_code: Optional[str] = None) -> dict[str, Any]:
    payload = dict(job.get("print_payload") or {})
    normalized_code = str(template_code or payload.get("template_code") or "").strip().lower()
    job_type = str(job.get("job_type") or "").strip().lower()
    template: dict[str, Any] = {}
    resolved_scope = _resolve_template_scope(job, payload)
    if normalized_code:
        direct_template = state.label_templates.get(normalized_code) or {}
        template_scope = str(payload.get("template_scope") or direct_template.get("template_scope") or resolved_scope or "").strip().lower()
        template = state.get_label_template(normalized_code, template_scope=template_scope) if template_scope else (direct_template or {})
    elif job_type == "bale_barcode_label":
        default_template_code = "wait_for_transtoshop" if resolved_scope == "warehouseout_bale" else "warehouse_in"
        template = state.get_label_template(default_template_code, template_scope=resolved_scope or "bale")
    if template:
        payload["template_code"] = template.get("template_code") or normalized_code
        payload["template_scope"] = template.get("template_scope") or payload.get("template_scope") or ("bale" if job_type == "bale_barcode_label" else "product")
        payload["template_fields"] = template.get("fields", payload.get("template_fields", []))
        payload["width_mm"] = int(template.get("width_mm") or payload.get("width_mm") or 60)
        payload["height_mm"] = int(template.get("height_mm") or payload.get("height_mm") or 40)
        payload["paper_preset"] = template.get("paper_preset") or payload.get("paper_preset") or ""
        payload["layout"] = dict(template.get("layout") or payload.get("layout") or {})
        payload["label_size"] = f"{payload['width_mm']}x{payload['height_mm']}"
    if not payload.get("label_size"):
        payload["label_size"] = str(job.get("label_size") or "60x40")
    if not payload.get("template_scope"):
        payload["template_scope"] = resolved_scope
    return payload


def _resolve_template_scope(job: dict[str, Any], payload: dict[str, Any]) -> str:
    explicit_scope = str(payload.get("template_scope") or "").strip().lower()
    if explicit_scope:
        return explicit_scope
    job_type = str(job.get("job_type") or "").strip().lower()
    if job_type == "bale_barcode_label":
        return "bale"
    if any(
        str(payload.get(key) or "").strip()
        for key in ["scan_token", "bale_barcode", "legacy_bale_barcode", "parcel_batch_no", "shipment_no", "supplier_name"]
    ):
        return "bale"
    if any(payload.get(key) not in {None, "", 0} for key in ["serial_no", "total_packages"]):
        return "bale"
    return "product"


def _is_bale_like_scope(template_scope: str) -> bool:
    normalized_scope = str(template_scope or "").strip().lower()
    return normalized_scope in {"bale", "warehouseout_bale"}


def _build_code128_svg(barcode_value: str, width_mm: float, height_mm: float) -> str:
    drawing = createBarcodeDrawing(
        "Code128",
        value=str(barcode_value or "-"),
        barHeight=max(height_mm * 0.32, 10) * mm,
        barWidth=0.26 * mm,
        humanReadable=False,
    )
    return drawing.asString("svg")


def _render_label_text(lines: list[str], max_chars: int = 26) -> str:
    wrapped: list[str] = []
    for line in lines:
        value = str(line or "").strip()
        if not value:
            continue
        wrapped.extend(wrap(value, max_chars) or [value])
    return "".join(f"<div class='label-line'>{escape(part)}</div>" for part in wrapped)


def _derive_bale_label_display_parts(payload: dict[str, Any]) -> dict[str, str]:
    encoded_barcode_value = str(
        payload.get("machine_code")
        or payload.get("barcode_value")
        or payload.get("scan_token")
        or "-"
    ).strip()
    identity_barcode_value = str(
        payload.get("legacy_bale_barcode")
        or payload.get("bale_barcode")
        or payload.get("barcode_value")
        or "-"
    ).strip()
    supplier_name = str(payload.get("supplier_name") or payload.get("product_name") or "").strip()
    category_display = str(payload.get("category_display") or "").strip() or " / ".join(
        [
            part
            for part in [str(payload.get("category_main") or "").strip(), str(payload.get("category_sub") or "").strip()]
            if part
        ]
    )
    package_position = str(
        payload.get("package_position_label")
        or payload.get("package_position")
        or f"第 {payload.get('serial_no', '-') or '-'} 包 / 共 {payload.get('total_packages', '-') or '-'} 包"
    ).strip()
    barcode_parts = [part.strip().upper() for part in identity_barcode_value.split("-") if part.strip()]
    primary_identity = supplier_name or category_display or identity_barcode_value
    package_compact = package_position.replace("PKG ", "").replace("/", "-").replace("第 ", "").replace(" 包 / 共 ", "-").replace(" 包", "") if package_position else ""
    shipment_date = ""
    if len(barcode_parts) >= 7 and barcode_parts[0] == "BALE" and barcode_parts[1] == "BL":
        shipment_date = barcode_parts[2]
        primary_identity = f"{barcode_parts[3]}-{barcode_parts[4]}"
        package_compact = f"{barcode_parts[5]}-{barcode_parts[6]}"
    bottom_human = str(
        payload.get("human_readable")
        or payload.get("machine_code")
        or payload.get("barcode_value")
        or payload.get("scan_token")
        or encoded_barcode_value
    ).strip()
    if not bottom_human:
        bottom_human = package_compact or (encoded_barcode_value[-18:] if len(encoded_barcode_value) > 18 else encoded_barcode_value)
    return {
        "barcode_value": encoded_barcode_value or "-",
        "bale_barcode": str(payload.get("bale_barcode") or encoded_barcode_value or "-").strip(),
        "legacy_bale_barcode": identity_barcode_value or "-",
        "supplier_name": supplier_name,
        "category_display": category_display,
        "package_position": package_position,
        "primary_identity": primary_identity or "-",
        "package_compact": package_compact or "-",
        "shipment_date": shipment_date,
        "bottom_human": bottom_human or "-",
    }


def _split_bale_category_parts(payload: dict[str, Any], display: dict[str, str]) -> tuple[str, str]:
    category_main = str(payload.get("category_main") or "").strip()
    category_sub = str(payload.get("category_sub") or "").strip()
    if category_main and category_sub:
        return category_main, category_sub
    display_parts = [
        part.strip()
        for part in str(display.get("category_display") or "").split("/")
        if part.strip()
    ]
    if not category_main and display_parts:
        category_main = display_parts[0]
    if not category_sub and len(display_parts) > 1:
        category_sub = " / ".join(display_parts[1:])
    return category_main, category_sub


def _compact_batch_trace_value(batch_no: str) -> str:
    normalized = str(batch_no or "").strip()
    if not normalized:
        return "-"
    date_match = re.search(r"20(\d{2})(\d{2})(\d{2})", normalized)
    serial_match = re.search(r"(\d{3,})$", normalized)
    compact_date = "".join(date_match.groups()) if date_match else ""
    compact_serial = serial_match.group(1)[-3:] if serial_match else ""
    compact = "-".join(part for part in [compact_date, compact_serial] if part)
    return compact or normalized


def _compact_inbound_time_value(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return "-"
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})[ T](\d{2}:\d{2})", normalized)
    if not match:
        return normalized
    return f"{match.group(2)}-{match.group(3)} {match.group(4)}"


def _get_bale_template_layout(payload: dict[str, Any], width_mm: float, height_mm: float) -> dict[str, Any]:
    return state._normalize_label_template_layout(
        "bale",
        payload.get("layout") if isinstance(payload.get("layout"), dict) else {},
        int(round(width_mm)),
        int(round(height_mm)),
    )


def _default_product_template_layout(width_mm: float, height_mm: float) -> dict[str, Any]:
    if width_mm <= 45:
        return {
            "paper_preset": f"{int(round(width_mm))}x{int(round(height_mm))}",
            "components": [
                {"id": "default_product_price", "type": "text", "enabled": True, "x_mm": 1.6, "y_mm": 1.8, "w_mm": 20.8, "h_mm": 7.6, "font_size": 20.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "price"},
                {"id": "default_product_short_suffix", "type": "text", "enabled": True, "x_mm": 24.2, "y_mm": 2.0, "w_mm": 12.8, "h_mm": 4.0, "font_size": 10.0, "font_weight": "700", "align": "center", "vertical_align": "middle", "render_mode": "bitmap", "content_source": "short_suffix"},
                {"id": "default_product_name", "type": "text", "enabled": True, "x_mm": 1.6, "y_mm": 10.2, "w_mm": 35.6, "h_mm": 4.6, "font_size": 10.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "product_name"},
                {"id": "default_product_divider", "type": "line", "enabled": True, "x_mm": 1.6, "y_mm": 16.0, "w_mm": 35.6, "h_mm": 0.5, "content_source": "none"},
                {"id": "default_product_barcode", "type": "barcode", "enabled": True, "x_mm": 1.6, "y_mm": 17.4, "w_mm": 35.6, "h_mm": 7.4, "align": "center", "content_source": "barcode_value"},
                {"id": "default_product_barcode_text", "type": "text", "enabled": True, "x_mm": 1.6, "y_mm": 25.6, "w_mm": 26.0, "h_mm": 2.0, "font_size": 8.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "barcode_value"},
            ],
        }
    return {
        "paper_preset": f"{int(round(width_mm))}x{int(round(height_mm))}",
        "components": [
            {"id": "default_product_price", "type": "text", "enabled": True, "x_mm": 2.4, "y_mm": 2.2, "w_mm": 33.0, "h_mm": 12.0, "font_size": 42.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "price"},
            {"id": "default_product_name", "type": "text", "enabled": True, "x_mm": 2.4, "y_mm": 15.0, "w_mm": 33.0, "h_mm": 6.2, "font_size": 22.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "product_name"},
            {"id": "default_product_short_suffix", "type": "text", "enabled": True, "x_mm": 38.0, "y_mm": 3.8, "w_mm": 18.0, "h_mm": 4.8, "font_size": 15.0, "font_weight": "700", "align": "center", "vertical_align": "middle", "render_mode": "bitmap", "content_source": "short_suffix"},
            {"id": "default_product_divider", "type": "line", "enabled": True, "x_mm": 2.4, "y_mm": 22.1, "w_mm": 55.2, "h_mm": 0.7, "content_source": "none"},
            {"id": "default_product_barcode", "type": "barcode", "enabled": True, "x_mm": 2.4, "y_mm": 24.0, "w_mm": 55.2, "h_mm": 9.8, "align": "center", "content_source": "barcode_value"},
            {"id": "default_product_barcode_text", "type": "text", "enabled": True, "x_mm": 2.4, "y_mm": 34.5, "w_mm": 40.0, "h_mm": 3.0, "font_size": 11.0, "font_weight": "700", "align": "left", "vertical_align": "top", "render_mode": "bitmap", "content_source": "barcode_value"},
        ],
    }


def _get_product_template_layout(payload: dict[str, Any], width_mm: float, height_mm: float) -> dict[str, Any]:
    layout = payload.get("layout") if isinstance(payload.get("layout"), dict) else {}
    normalized = state._normalize_label_template_layout(
        "product",
        layout,
        int(round(width_mm)),
        int(round(height_mm)),
    )
    if isinstance(normalized.get("components"), list) and normalized.get("components"):
        return normalized
    return _default_product_template_layout(width_mm, height_mm)


def _format_product_price(value: Any) -> str:
    if value in {None, ""}:
        return ""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if numeric <= 0:
        return ""
    if abs(numeric - round(numeric)) < 0.01:
        return f"KES {int(round(numeric))}"
    return f"KES {numeric:.2f}"


def _derive_product_label_display_parts(payload: dict[str, Any]) -> dict[str, str]:
    barcode_value = str(payload.get("barcode_value") or payload.get("human_readable") or "-").strip() or "-"
    product_name = str(payload.get("product_name") or payload.get("category_sub") or payload.get("category_main") or "").strip() or "-"
    short_suffix = str(payload.get("short_suffix") or payload.get("rack_code") or "").strip().upper()
    return {
        "price": _format_product_price(payload.get("price") or payload.get("selling_price_kes") or payload.get("launch_price")),
        "product_name": product_name,
        "short_suffix": short_suffix,
        "barcode_value": barcode_value,
    }


def _build_product_template_content_map(payload: dict[str, Any], display: dict[str, str]) -> dict[str, str]:
    return {
        "price": str(display.get("price") or "").strip(),
        "product_name": str(display.get("product_name") or "").strip(),
        "short_suffix": str(display.get("short_suffix") or "").strip(),
        "barcode_value": str(display.get("barcode_value") or "").strip(),
        "none": "",
    }


def _build_bale_template_content_map(payload: dict[str, Any], display: dict[str, str]) -> dict[str, str]:
    supplier_name = str(display.get("supplier_name") or "").strip()
    category_display = str(display.get("category_display") or "").strip()
    category_main, category_sub = _split_bale_category_parts(payload, display)
    package_position = str(display.get("package_position") or "").strip()
    compact_package = str(display.get("package_compact") or "").strip()
    compact_package = "" if compact_package == "-" else compact_package
    package_line = compact_package or package_position
    headline_category_package = "\n".join(
        part for part in [
            supplier_name,
            " · ".join(part for part in [category_display, package_line] if part),
        ] if part
    ).strip()
    headline_supplier_package = "\n".join(part for part in [supplier_name, package_line] if part).strip()
    headline_category_only = "\n".join(part for part in [category_display, package_line] if part).strip()
    shipment_no = str(payload.get("shipment_no") or "").strip()
    parcel_batch_no = str(payload.get("parcel_batch_no") or "").strip()
    received_at = str(
        payload.get("unload_date")
        or payload.get("received_at")
        or payload.get("created_at")
        or ""
    ).strip()
    shipment_batch = " / ".join(part for part in [shipment_no, parcel_batch_no] if part).strip()
    serial_no = int(payload.get("serial_no") or 0)
    total_packages = int(payload.get("total_packages") or 0)
    piece_current = str(serial_no) if serial_no > 0 else "-"
    piece_total = str(total_packages) if total_packages > 0 else "-"
    trace_code = str(display.get("barcode_value") or "").strip() or "-"
    trace_batch = _compact_batch_trace_value(parcel_batch_no)
    template_code = str(payload.get("template_code") or "").strip().lower()
    is_warehouse_in_label = template_code in {"warehouse_in", "warehouse_in_60x40"}
    display_code = str(
        payload.get("display_code")
        or payload.get("bale_barcode")
        or display.get("bale_barcode")
        or ""
    ).strip().upper()
    trace_shipment = shipment_no or "-"
    trace_inbound = _compact_inbound_time_value(received_at)
    store_name = str(payload.get("store_name") or payload.get("store_display") or "").strip()
    transfer_order_no = str(payload.get("transfer_order_no") or "").strip()
    bale_piece_summary = str(payload.get("bale_piece_summary") or payload.get("package_position_label") or "").strip()
    total_quantity = str(payload.get("total_quantity") or "").strip()
    packing_list = str(payload.get("packing_list") or "").strip()
    dispatch_bale_no = str(
        payload.get("machine_code")
        or payload.get("barcode_value")
        or payload.get("scan_token")
        or payload.get("dispatch_bale_no")
        or ""
    ).strip().upper()
    outbound_time = str(payload.get("outbound_time") or "").strip()
    status = str(payload.get("status") or "").strip()
    cat = str(payload.get("cat") or "").strip()
    sub = str(payload.get("sub") or "").strip()
    grade = str(payload.get("grade") or "").strip()
    qty = str(payload.get("qty") or "").strip()
    weight = str(payload.get("weight") or "").strip()
    code = str(payload.get("code") or display_code or dispatch_bale_no or trace_code).strip().upper()
    return {
        "supplier_category_package": headline_category_package,
        "supplier_package": headline_supplier_package,
        "category_package": headline_category_only,
        "supplier_category": "\n".join(part for part in [supplier_name, category_display] if part).strip(),
        "package_only": package_line or package_position,
        "scan_token": str(display.get("barcode_value") or "").strip(),
        "shipment_no": shipment_no,
        "parcel_batch_no": parcel_batch_no,
        "shipment_batch": shipment_batch,
        "bale_barcode": str(display.get("bale_barcode") or "").strip(),
        "top_supplier": f"SUP: {supplier_name or '-'}",
        "top_major": f"CAT: {category_main or category_display or '-'}",
        "top_minor": f"SUB: {category_sub or '-'}",
        "piece_current": f"No: {piece_current}",
        "piece_total": f"Total: {piece_total}",
        "trace_code": f"{'Machine' if is_warehouse_in_label else 'Code'}: {trace_code}",
        "trace_batch": f"Display: {display_code or '-'}" if is_warehouse_in_label else f"Batch: {trace_batch}",
        "trace_shipment": f"Encoded: {trace_code}" if is_warehouse_in_label else f"Ship: {trace_shipment}",
        "trace_inbound": f"In: {trace_inbound}",
        "store_name": store_name,
        "transfer_order_no": transfer_order_no,
        "bale_piece_summary": bale_piece_summary,
        "total_quantity": f"Total: {total_quantity}" if total_quantity else "",
        "packing_list": packing_list,
        "dispatch_bale_no": dispatch_bale_no,
        "outbound_time": f"Out: {outbound_time}" if outbound_time else "",
        "status": f"STATUS: {status}" if status else "",
        "cat": f"CAT: {cat}" if cat else "",
        "sub": f"SUB: {sub}" if sub else "",
        "grade": f"GRADE: {grade}" if grade else "",
        "qty": f"QTY: {qty}" if qty else "",
        "weight": f"WEIGHT: {weight}" if weight else "",
        "code": f"CODE: {code}" if code else "",
        "none": "",
    }


def _render_bale_template_component_html(component: dict[str, Any], content_value: str, barcode_svg: str) -> str:
    left = float(component.get("x_mm") or 0)
    top = float(component.get("y_mm") or 0)
    width = float(component.get("w_mm") or 0)
    height = float(component.get("h_mm") or 0)
    base_style = (
        f"left:{left}mm;top:{top}mm;width:{width}mm;height:{height}mm;"
        "position:absolute;box-sizing:border-box;overflow:hidden;"
    )
    component_type = str(component.get("type") or "").strip().lower()
    if component_type == "barcode":
        return (
            f"<div class='label-block label-barcode align-{escape(str(component.get('align') or 'center'))}' "
            f"style=\"{base_style}\">{barcode_svg}</div>"
        )
    if component_type == "line":
        return f"<div class='label-block label-line' style=\"{base_style}background:#111;\"></div>"
    font_size = _fit_text_component_font_size(component, content_value)
    font_weight = "700" if str(component.get("font_weight") or "") == "700" else "400"
    text_html = escape(str(content_value or "")).replace("\n", "<br />")
    return (
        f"<div class='label-block label-text align-{escape(str(component.get('align') or 'left'))} "
        f"valign-{escape(str(component.get('vertical_align') or 'top'))}' "
        f"style=\"{base_style}font-size:{font_size}px;font-weight:{font_weight};\">{text_html}</div>"
    )


def _build_barcode_preview_html(
    job: dict[str, Any],
    payload: dict[str, Any],
    *,
    autoprint: bool = False,
) -> str:
    width_mm, height_mm = _resolve_label_dimensions(payload, str(job.get("label_size") or "60x40"))
    template_scope = _resolve_template_scope(job, payload)
    if _is_bale_like_scope(template_scope):
        display = _derive_bale_label_display_parts(payload)
        layout = _get_bale_template_layout(payload, width_mm, height_mm)
        value_map = _build_bale_template_content_map(payload, display)
        barcode_value = display["barcode_value"]
    else:
        display = _derive_product_label_display_parts(payload)
        layout = _get_product_template_layout(payload, width_mm, height_mm)
        value_map = _build_product_template_content_map(payload, display)
        barcode_value = display["barcode_value"]
    barcode_svg = _build_code128_svg(barcode_value, width_mm, height_mm)
    component_html = "".join(
        _render_bale_template_component_html(
            component,
            value_map.get(str(component.get("content_source") or "").strip(), ""),
            barcode_svg,
        )
        for component in layout.get("components", [])
        if component.get("enabled")
    )
    auto_print_script = (
        "<script>window.addEventListener('load',()=>setTimeout(()=>window.print(),180));</script>"
        if autoprint
        else ""
    )
    return f"""
    <html>
      <head>
        <style>
          @page {{ size: {width_mm}mm {height_mm}mm; margin: 0; }}
          html, body {{
            margin: 0;
            padding: 0;
            width: 100%;
            height: 100%;
            background: #ffffff;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          }}
          body {{
            display: flex;
            align-items: center;
            justify-content: center;
          }}
          .label {{
            width: {width_mm}mm;
            height: {height_mm}mm;
            box-sizing: border-box;
            border: 1px solid #111;
            position: relative;
            overflow: hidden;
            background:
              linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(250,247,241,0.98) 100%);
          }}
          .label-block {{
            position: absolute;
            overflow: hidden;
            box-sizing: border-box;
          }}
          .label-text {{
            display: flex;
            flex-direction: column;
            line-height: 1.15;
            color: #111;
            word-break: break-word;
            white-space: pre-line;
          }}
          .label-line {{
            background: #111;
          }}
          .label-barcode {{
            display: flex;
            align-items: center;
            justify-content: center;
          }}
          .label-barcode svg {{
            width: 100%;
            height: 100%;
          }}
          .align-left {{ text-align: left; }}
          .align-center {{ text-align: center; }}
          .align-right {{ text-align: right; }}
          .valign-top {{ justify-content: flex-start; }}
          .valign-middle {{ justify-content: center; }}
          .valign-bottom {{ justify-content: flex-end; }}
        </style>
      </head>
      <body>
        <div class="label">{component_html}</div>
        {auto_print_script}
      </body>
    </html>
    """


def _draw_wrapped_text(
    pdf: canvas.Canvas,
    text: str,
    *,
    x: float,
    y: float,
    max_chars: int,
    font_name: str,
    font_size: float,
    leading: float,
) -> float:
    current_y = y
    for line in wrap(str(text or "").strip(), max_chars) or [str(text or "").strip()]:
        if not line:
            continue
        pdf.setFont(font_name, font_size)
        pdf.drawString(x, current_y, line)
        current_y -= leading
    return current_y


def _render_barcode_print_job_pdf(job: dict[str, Any], payload: dict[str, Any], output_path: str) -> None:
    width_mm, height_mm = _resolve_label_dimensions(payload, str(job.get("label_size") or "60x40"))
    page_width = width_mm * mm
    page_height = height_mm * mm
    margin = 2.4 * mm
    pdf = canvas.Canvas(output_path, pagesize=(page_width, page_height))
    pdf.setStrokeColorRGB(0.07, 0.07, 0.07)
    pdf.rect(0.6 * mm, 0.6 * mm, page_width - 1.2 * mm, page_height - 1.2 * mm)

    display = _derive_bale_label_display_parts(payload)
    barcode_value = display["barcode_value"] or "-"

    current_y = page_height - margin - 0.5 * mm
    pdf.setFont("Helvetica-Bold", 9)
    current_y = _draw_wrapped_text(
        pdf,
        display["primary_identity"],
        x=margin,
        y=current_y,
        max_chars=22 if width_mm >= 60 else 18,
        font_name="Helvetica-Bold",
        font_size=9 if width_mm >= 60 else 7,
        leading=9 if width_mm >= 60 else 7.5,
    ) - 0.8 * mm

    compact_meta = " · ".join(part for part in [display["package_compact"], display["shipment_date"]] if part)
    if compact_meta:
        current_y = _draw_wrapped_text(
            pdf,
            compact_meta,
            x=margin,
            y=current_y,
            max_chars=28 if width_mm >= 60 else 22,
            font_name="Helvetica",
            font_size=7 if width_mm >= 60 else 6,
            leading=7,
        ) - 0.8 * mm

    barcode_drawing = createBarcodeDrawing(
        "Code128",
        value=barcode_value,
        barHeight=max(height_mm * 0.22, 9) * mm,
        barWidth=0.26 * mm,
        humanReadable=False,
    )
    available_width = page_width - 2 * margin
    available_height = max(page_height * 0.34, 12 * mm)
    scale_x = available_width / max(float(barcode_drawing.width), 1.0)
    scale_y = available_height / max(float(barcode_drawing.height), 1.0)
    scale = min(scale_x, scale_y, 1.0)
    draw_width = float(barcode_drawing.width) * scale
    draw_height = float(barcode_drawing.height) * scale
    barcode_y = max(current_y - draw_height, 9 * mm)
    pdf.saveState()
    pdf.translate((page_width - draw_width) / 2, barcode_y)
    pdf.scale(scale, scale)
    renderPDF.draw(barcode_drawing, pdf, 0, 0)
    pdf.restoreState()

    text_y = barcode_y - 3.6 * mm
    pdf.setFont("Helvetica-Bold", 7.2 if width_mm >= 60 else 6.2)
    pdf.drawString(margin, text_y, display["bottom_human"])
    text_y -= 3.4 * mm
    if display["primary_identity"]:
        text_y = _draw_wrapped_text(
            pdf,
            display["primary_identity"],
            x=margin,
            y=text_y,
            max_chars=28 if width_mm >= 60 else 22,
            font_name="Helvetica",
            font_size=6.2 if width_mm >= 60 else 5.8,
            leading=6.5,
        ) - 0.4 * mm
    if compact_meta:
        _draw_wrapped_text(
            pdf,
            compact_meta,
            x=margin,
            y=text_y,
            max_chars=28 if width_mm >= 60 else 22,
            font_name="Helvetica",
            font_size=6.1 if width_mm >= 60 else 5.7,
            leading=6.2,
        )

    pdf.showPage()
    pdf.save()


def _mm_to_dots(value_mm: float, dpi: int = 203) -> int:
    return max(int(round((float(value_mm) / 25.4) * dpi)), 1)


def _wrap_tspl_text(value: str, max_chars: int = 26, max_lines: int = 3) -> list[str]:
    lines: list[str] = []
    for raw_line in wrap(str(value or "").strip(), max_chars) or [str(value or "").strip()]:
        trimmed = raw_line.strip()
        if not trimmed:
            continue
        lines.append(trimmed)
        if len(lines) >= max_lines:
            break
    return lines


def _escape_tspl_text(value: str) -> str:
    text = str(value or "").replace("\\", " ").replace("\"", "'").replace("\n", " ").replace("\r", " ")
    return "".join(ch for ch in text if 32 <= ord(ch) <= 126)


def _dots_to_mm(value_dots: int, dpi: int = 203) -> float:
    return (int(value_dots or 0) / dpi) * 25.4


def _estimate_code128_width_dots(barcode_value: str, narrow_dots: int) -> int:
    normalized = str(barcode_value or "").strip()
    if not normalized:
        return 0
    return max((len(normalized) * 11 + 35) * max(int(narrow_dots or 1), 1), 120)


def _select_tspl_barcode_widths(barcode_value: str, available_width_dots: int) -> tuple[int, int]:
    width_dots = max(int(available_width_dots or 0), 0)
    quiet_margin_dots = max(int(round(width_dots * 0.12)), 24)
    for narrow_dots, wide_dots in ((2, 4), (1, 2)):
        estimated_width = _estimate_code128_width_dots(barcode_value, narrow_dots)
        if estimated_width + quiet_margin_dots <= width_dots:
            return narrow_dots, wide_dots
    return 1, 2


def _resolve_tspl_barcode_origin(x_dots: int, width_dots: int, barcode_value: str, narrow_dots: int) -> int:
    left = max(int(x_dots or 0), 0)
    available_width = max(int(width_dots or 0), 0)
    quiet_zone = max(int(round(available_width * 0.04)), 10)
    estimated_width = _estimate_code128_width_dots(barcode_value, narrow_dots)
    max_origin = left + max(available_width - estimated_width - quiet_zone, 0)
    return min(left + quiet_zone, max_origin)


def _estimate_qr_module_count(qr_value: str) -> int:
    length = len(str(qr_value or "").strip())
    if length <= 25:
        return 21
    if length <= 47:
        return 25
    if length <= 77:
        return 29
    if length <= 114:
        return 33
    if length <= 154:
        return 37
    return 41


def _select_tspl_qr_cell_width(qr_value: str, available_width_dots: int, available_height_dots: int) -> tuple[int, int]:
    modules = _estimate_qr_module_count(qr_value)
    usable_dots = max(min(int(available_width_dots or 0), int(available_height_dots or 0)) - 16, 32)
    cell_width = max(min(usable_dots // max(modules, 1), 8), 2)
    return cell_width, modules


def _resolve_tspl_qr_origin(
    x_dots: int,
    y_dots: int,
    width_dots: int,
    height_dots: int,
    modules: int,
    cell_width: int,
) -> tuple[int, int]:
    qr_size = max(int(modules or 21) * max(int(cell_width or 2), 1), 0)
    left = max(int(x_dots or 0), 0)
    top = max(int(y_dots or 0), 0)
    offset_x = max((max(int(width_dots or 0), 0) - qr_size) // 2, 0)
    offset_y = max((max(int(height_dots or 0), 0) - qr_size) // 2, 0)
    return left + offset_x, top + offset_y


def _tspl_text_font_spec(font_size: float) -> tuple[str, int, int, int]:
    normalized = max(float(font_size or 6), 5.0)
    if normalized >= 14:
        return "3", 2, 2, 28
    if normalized >= 10:
        return "3", 1, 1, 22
    if normalized >= 7:
        return "2", 1, 1, 18
    return "1", 1, 1, 16


def _estimate_tspl_line_width(text: str, font_size: float, font_weight: str) -> int:
    normalized = str(text or "").strip()
    if not normalized:
        return 0
    per_char = max(int(round(float(font_size or 6) * (2.3 if font_weight == "700" else 2.0))), 10)
    return len(normalized) * per_char


def _estimate_wrapped_line_count(text: str, max_chars: int) -> int:
    normalized_limit = max(int(max_chars or 1), 1)
    total_lines = 0
    for raw_line in str(text or "").splitlines() or [str(text or "").strip()]:
        trimmed = str(raw_line or "").strip()
        if not trimmed:
            continue
        total_lines += max(int(math.ceil(len(trimmed) / normalized_limit)), 1)
    return total_lines or 1


def _fit_text_component_font_size(component: dict[str, Any], content_value: str, minimum_font_size: float = 5.0) -> float:
    text = str(content_value or "").strip()
    preferred_size = max(float(component.get("font_size") or 6), minimum_font_size)
    if not text:
        return preferred_size
    font_weight = str(component.get("font_weight") or "400").strip()
    width_mm = max(float(component.get("w_mm") or 10), 1.0)
    height_mm = max(float(component.get("h_mm") or 4), 1.0)
    width_dots = _mm_to_dots(width_mm)
    current_size = preferred_size
    while current_size >= minimum_font_size:
        _, _, _, line_height_dots = _tspl_text_font_spec(current_size)
        chars_per_line = max(
            int(width_dots / max(_estimate_tspl_line_width("W", current_size, font_weight), 10)),
            1,
        )
        wrapped_lines = _estimate_wrapped_line_count(text, chars_per_line)
        total_height_mm = _dots_to_mm(line_height_dots * wrapped_lines)
        if total_height_mm <= height_mm + 0.35:
            return round(current_size, 1)
        current_size -= 0.5
    return minimum_font_size


def _append_tspl_text_component(tspl_lines: list[str], component: dict[str, Any], content_value: str) -> None:
    text = str(content_value or "").strip()
    if not text:
        return
    font_size = _fit_text_component_font_size(component, text)
    font_weight = str(component.get("font_weight") or "400").strip()
    font_code, x_mul, y_mul, line_height = _tspl_text_font_spec(font_size)
    x_dots = _mm_to_dots(component.get("x_mm") or 0)
    y_dots = _mm_to_dots(component.get("y_mm") or 0)
    width_dots = _mm_to_dots(component.get("w_mm") or 10)
    height_dots = _mm_to_dots(component.get("h_mm") or 4)
    max_chars = max(int(width_dots / max(_estimate_tspl_line_width("W", font_size, font_weight), 10)), 4)
    max_lines = max(int(height_dots / max(line_height, 14)), 1)
    wrapped_lines = _wrap_tspl_text(text, max_chars=max_chars, max_lines=max_lines)
    for index, line in enumerate(wrapped_lines):
        line_text = _escape_tspl_text(line)
        if not line_text:
            continue
        line_width = _estimate_tspl_line_width(line_text, font_size, font_weight)
        align = str(component.get("align") or "left").strip().lower()
        line_x = x_dots
        if align == "center":
            line_x = x_dots + max(int((width_dots - line_width) / 2), 0)
        elif align == "right":
            line_x = x_dots + max(width_dots - line_width, 0)
        line_y = y_dots + index * line_height
        tspl_lines.append(f'TEXT {line_x},{line_y},"{font_code}",0,{x_mul},{y_mul},"{line_text}"')


def _resolve_candidate_bitmap_font_path() -> str:
    for path in _CANDIDATE_BITMAP_FONT_PATHS:
        if os.path.exists(path):
            return path
    return _CANDIDATE_BITMAP_FONT_PATHS[0]


def _measure_bitmap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont) -> int:
    if not text:
        return 0
    bbox = draw.textbbox((0, 0), text, font=font)
    return max(int(bbox[2] - bbox[0]), 0)


def _wrap_bitmap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont, max_width: int) -> list[str]:
    normalized = str(text or "").strip()
    if not normalized:
        return []
    lines: list[str] = []
    for raw_line in normalized.splitlines() or [normalized]:
        source = str(raw_line or "").strip()
        if not source:
            continue
        current = ""
        for char in source:
            trial = f"{current}{char}"
            if current and _measure_bitmap_text(draw, trial, font) > max_width:
                lines.append(current)
                current = char
            else:
                current = trial
        if current:
            lines.append(current)
    return lines or [normalized]


def _render_candidate_bitmap_text(
    component: dict[str, Any],
    content_value: str,
) -> tuple[int, int, int, int, bytes]:
    text = str(content_value or "").strip()
    if not text:
        return 0, 0, 0, 0, b""

    width_dots = max(_mm_to_dots(component.get("w_mm") or 10), 8)
    height_dots = max(_mm_to_dots(component.get("h_mm") or 4), 8)
    x_dots = _mm_to_dots(component.get("x_mm") or 0)
    y_dots = _mm_to_dots(component.get("y_mm") or 0)
    preferred_px = max(int(round(float(component.get("font_size") or 8) * 2.0)), 10)
    min_px = 9
    font_path = _resolve_candidate_bitmap_font_path()
    scratch = Image.new("L", (8, 8), 255)
    scratch_draw = ImageDraw.Draw(scratch)

    best_font: Optional[ImageFont.FreeTypeFont] = None
    best_lines: list[str] = []
    best_line_height = 0

    start_px = min(preferred_px, max(height_dots - 2, min_px))
    for font_px in range(start_px, min_px - 1, -1):
        font = ImageFont.truetype(font_path, font_px)
        lines = _wrap_bitmap_text(scratch_draw, text, font, max(width_dots - 2, 6))
        ascent, descent = font.getmetrics()
        line_height = max(ascent + descent + 1, font_px)
        total_height = line_height * len(lines)
        if total_height <= height_dots - 1:
            best_font = font
            best_lines = lines
            best_line_height = line_height
            break

    if best_font is None:
        best_font = ImageFont.truetype(font_path, min_px)
        best_lines = _wrap_bitmap_text(scratch_draw, text, best_font, max(width_dots - 2, 6))
        ascent, descent = best_font.getmetrics()
        best_line_height = max(ascent + descent + 1, min_px)

    image = Image.new("L", (width_dots, height_dots), 255)
    draw = ImageDraw.Draw(image)
    total_height = best_line_height * len(best_lines)
    vertical_align = str(component.get("vertical_align") or "top").strip().lower()
    if vertical_align == "middle":
        current_y = max(int((height_dots - total_height) / 2), 0)
    elif vertical_align == "bottom":
        current_y = max(height_dots - total_height, 0)
    else:
        current_y = 0
    align = str(component.get("align") or "left").strip().lower()
    for line in best_lines:
        line_width = _measure_bitmap_text(draw, line, best_font)
        line_x = 0
        if align == "center":
            line_x = max(int((width_dots - line_width) / 2), 0)
        elif align == "right":
            line_x = max(width_dots - line_width, 0)
        draw.text((line_x, current_y), line, font=best_font, fill=0)
        current_y += best_line_height

    image = image.point(lambda p: 0 if p < 180 else 255, mode="1")
    width_bytes = int(math.ceil(width_dots / 8))
    bitmap = bytearray()
    pixels = image.load()
    for row in range(height_dots):
        for byte_index in range(width_bytes):
            value = 0
            for bit in range(8):
                pixel_x = byte_index * 8 + bit
                if pixel_x >= width_dots:
                    continue
                # DL-720C's TSPL BITMAP path treats cleared bits as inked dots.
                # Encode white background as 1-bits so text prints black on white.
                if pixels[pixel_x, row] != 0:
                    value |= 1 << (7 - bit)
            bitmap.append(value)
    return x_dots, y_dots, width_bytes, height_dots, bytes(bitmap)


def _build_candidate_lab_tspl_payload(payload: dict[str, Any]) -> bytes:
    width_mm = max(float(payload.get("width_mm") or 60), 20.0)
    height_mm = max(float(payload.get("height_mm") or 40), 20.0)
    width_dots = _mm_to_dots(width_mm)
    height_dots = _mm_to_dots(height_mm)
    component_rows = payload.get("blocks") if isinstance(payload.get("blocks"), list) else []

    tspl_lines = [
        f"SIZE {int(round(width_mm))} mm,{int(round(height_mm))} mm",
        "GAP 2 mm,0 mm",
        "DENSITY 8",
        "SPEED 4",
        "DIRECTION 1",
        "REFERENCE 0,0",
        "CLS",
    ]
    raw_tspl = bytearray(("\n".join(tspl_lines) + "\n").encode("ascii"))

    for block in component_rows:
        block_type = str(block.get("type") or "").strip().lower()
        x_mm = float(block.get("x_mm") or 0)
        y_mm = float(block.get("y_mm") or 0)
        raw_w_mm = float(block.get("w_mm") or 1)
        raw_h_mm = float(block.get("h_mm") or 1)
        if block_type == "line":
            w_mm = max(raw_w_mm, 0.1)
            h_mm = max(raw_h_mm, 0.1)
        else:
            w_mm = max(raw_w_mm, 1.0)
            h_mm = max(raw_h_mm, 1.0)
        x_dots = _mm_to_dots(x_mm)
        y_dots = _mm_to_dots(y_mm)
        block_width_dots = _mm_to_dots(w_mm)
        block_height_dots = _mm_to_dots(h_mm)
        if block_type == "barcode":
            barcode_value = _escape_tspl_text(str(block.get("value") or "").strip())
            if not barcode_value:
                continue
            barcode_height = min(block_height_dots, max(height_dots - y_dots - 12, 48))
            narrow_dots, wide_dots = _select_tspl_barcode_widths(barcode_value, block_width_dots)
            barcode_x = _resolve_tspl_barcode_origin(x_dots, block_width_dots, barcode_value, narrow_dots)
            raw_tspl.extend(
                f'BARCODE {barcode_x},{y_dots},"128",{barcode_height},0,0,{narrow_dots},{wide_dots},"{barcode_value}"\n'.encode("ascii")
            )
            continue
        if block_type == "qr":
            qr_value = _escape_tspl_text(str(block.get("value") or "").strip())
            if not qr_value:
                continue
            cell_width, modules = _select_tspl_qr_cell_width(qr_value, block_width_dots, block_height_dots)
            qr_x, qr_y = _resolve_tspl_qr_origin(x_dots, y_dots, block_width_dots, block_height_dots, modules, cell_width)
            raw_tspl.extend(
                f'QRCODE {qr_x},{qr_y},Q,{cell_width},A,0,M2,S7,"{qr_value}"\n'.encode("ascii")
            )
            continue
        if block_type == "line":
            raw_tspl.extend(
                f"BAR {x_dots},{y_dots},{max(block_width_dots, 1)},{max(block_height_dots, 1)}\n".encode("ascii")
            )
            continue
        if block_type == "badgerow":
            value = " · ".join(
                part for part in [str(item).strip() for item in block.get("values", [])] if part
            )
        else:
            value = str(block.get("value") or "").strip()
        if not value:
            continue
        component = {
            "x_mm": x_mm,
            "y_mm": y_mm,
            "w_mm": w_mm,
            "h_mm": h_mm,
            "font_size": float(block.get("font_size") or 8),
            "font_weight": "700" if str(block.get("font_weight") or "700").strip() == "700" else "400",
            "align": str(block.get("align") or "left").strip().lower(),
            "vertical_align": str(block.get("vertical_align") or "top").strip().lower(),
        }
        bitmap_x, bitmap_y, bitmap_width_bytes, bitmap_height, bitmap_payload = _render_candidate_bitmap_text(component, value)
        if bitmap_payload:
            raw_tspl.extend(f"BITMAP {bitmap_x},{bitmap_y},{bitmap_width_bytes},{bitmap_height},0,".encode("ascii"))
            raw_tspl.extend(bitmap_payload)
            raw_tspl.extend(b"\n")
    raw_tspl.extend(b"PRINT 1,1\n")
    return bytes(raw_tspl)


def _build_candidate_lab_tspl_batch_payload(payload: dict[str, Any]) -> bytes:
    candidates = payload.get("candidates") if isinstance(payload.get("candidates"), list) else []
    return b"".join(
        _build_candidate_lab_tspl_payload(
            {
                "width_mm": payload.get("width_mm") or 60,
                "height_mm": payload.get("height_mm") or 40,
                "blocks": candidate.get("blocks", []),
            }
        )
        for candidate in candidates
    )


def _is_tspl_raw_printer(printer_row: Optional[dict[str, Any]]) -> bool:
    if not printer_row:
        return False
    fingerprint = " ".join(
        [
            str(printer_row.get("name") or ""),
            str(printer_row.get("device_uri") or ""),
        ]
    ).lower()
    return "deli" in fingerprint or "dl-720" in fingerprint or "dl720" in fingerprint


def _append_tspl_bitmap_text_component(raw_tspl: bytearray, component: dict[str, Any], content_value: str) -> None:
    bitmap_x, bitmap_y, bitmap_width_bytes, bitmap_height, bitmap_payload = _render_candidate_bitmap_text(component, content_value)
    if not bitmap_payload:
        return
    raw_tspl.extend(f"BITMAP {bitmap_x},{bitmap_y},{bitmap_width_bytes},{bitmap_height},0,".encode("ascii"))
    raw_tspl.extend(bitmap_payload)
    raw_tspl.extend(_TSPL_EOL)


def _build_tspl_barcode_payload(job: dict[str, Any], payload: dict[str, Any], *, home_before_print: bool = False) -> bytes:
    width_mm, height_mm = _resolve_label_dimensions(payload, str(job.get("label_size") or "60x40"))
    height_dots = _mm_to_dots(height_mm)
    template_scope = _resolve_template_scope(job, payload)
    if _is_bale_like_scope(template_scope):
        display = _derive_bale_label_display_parts(payload)
        layout = _get_bale_template_layout(payload, width_mm, height_mm)
        value_map = _build_bale_template_content_map(payload, display)
        barcode_value = _escape_tspl_text(display["barcode_value"])
    else:
        display = _derive_product_label_display_parts(payload)
        layout = _get_product_template_layout(payload, width_mm, height_mm)
        value_map = _build_product_template_content_map(payload, display)
        barcode_value = _escape_tspl_text(display["barcode_value"])

    tspl_lines = [
        f"SIZE {int(round(width_mm))} mm,{int(round(height_mm))} mm",
        "GAP 2 mm,0 mm",
        "DENSITY 8",
        "SPEED 4",
        "DIRECTION 1",
        "REFERENCE 0,0",
        "SET PEEL OFF",
        "SET TEAR OFF",
    ]
    if home_before_print:
        tspl_lines.append("HOME")
    tspl_lines.append("CLS")
    raw_tspl = bytearray(("\r\n".join(tspl_lines) + "\r\n").encode("ascii"))

    for component in layout.get("components", []):
        if not component.get("enabled"):
            continue
        component_type = str(component.get("type") or "").strip().lower()
        if component_type == "barcode":
            x_dots = _mm_to_dots(component.get("x_mm") or 0)
            y_dots = _mm_to_dots(component.get("y_mm") or 0)
            width_dots = _mm_to_dots(component.get("w_mm") or width_mm)
            barcode_height = min(_mm_to_dots(component.get("h_mm") or 12), max(height_dots - y_dots - 12, 48))
            narrow_dots, wide_dots = _select_tspl_barcode_widths(barcode_value, width_dots)
            barcode_x = _resolve_tspl_barcode_origin(x_dots, width_dots, barcode_value, narrow_dots)
            raw_tspl.extend(
                f'BARCODE {barcode_x},{y_dots},"128",{barcode_height},0,0,{narrow_dots},{wide_dots},"{barcode_value}"\r\n'.encode("ascii")
            )
            continue
        if component_type == "line":
            x_dots = _mm_to_dots(component.get("x_mm") or 0)
            y_dots = _mm_to_dots(component.get("y_mm") or 0)
            width_dots = _mm_to_dots(max(float(component.get("w_mm") or 0.1), 0.1))
            height_dots_component = _mm_to_dots(max(float(component.get("h_mm") or 0.1), 0.1))
            raw_tspl.extend(
                f"BAR {x_dots},{y_dots},{max(width_dots, 1)},{max(height_dots_component, 1)}\r\n".encode("ascii")
            )
            continue
        source = str(component.get("content_source") or "").strip()
        content_value = value_map.get(source, "")
        if str(component.get("render_mode") or "").strip().lower() == "bitmap":
            _append_tspl_bitmap_text_component(raw_tspl, component, content_value)
        else:
            text_lines: list[str] = []
            _append_tspl_text_component(text_lines, component, content_value)
            if text_lines:
                raw_tspl.extend(("\r\n".join(text_lines) + "\r\n").encode("ascii"))
    raw_tspl.extend(b"PRINT 1,1\r\n")
    return bytes(raw_tspl)


def _build_tspl_barcode_batch_payload(jobs: list[dict[str, Any]], payloads: list[dict[str, Any]]) -> bytes:
    raw_chunks: list[bytes] = []
    for index, payload in enumerate(payloads):
        job = jobs[index] if index < len(jobs) else jobs[-1]
        copies = max(int(payload.get("copies") or job.get("copies") or 1), 1)
        single_payload = _build_tspl_barcode_payload(job, payload, home_before_print=False)
        raw_chunks.append(single_payload * copies)
    return b"".join(raw_chunks)


def _build_print_debug_snapshot(
    printer_name: str,
    payload: bytes,
    *,
    copies: int,
    print_result: Optional[subprocess.CompletedProcess[str]] = None,
    error: Optional[str] = None,
) -> dict[str, Any]:
    created_at = datetime.now(timezone.utc).isoformat()
    printer_stub = re.sub(r"[^a-z0-9]+", "_", str(printer_name or "").strip().lower()).strip("_") or "printer"
    file_stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    debug_dir = Path(settings.data_dir) / "print_debug"
    debug_dir.mkdir(parents=True, exist_ok=True)

    base_name = f"{file_stamp}_{printer_stub}"
    tspl_path = debug_dir / f"{base_name}.tspl"
    meta_path = debug_dir / f"{base_name}.json"
    latest_tspl_path = debug_dir / f"latest_{printer_stub}.tspl"
    latest_meta_path = debug_dir / f"latest_{printer_stub}.json"

    tspl_path.write_bytes(payload)
    latest_tspl_path.write_bytes(payload)

    metadata = {
        "created_at": created_at,
        "printer_name": printer_name,
        "copies": max(int(copies or 1), 1),
        "payload_bytes": len(payload),
        "payload_sha256": hashlib.sha256(payload).hexdigest(),
        "tspl_path": str(tspl_path),
        "latest_tspl_path": str(latest_tspl_path),
        "error": str(error or "").strip(),
    }
    if print_result is not None:
        metadata.update(
            {
                "lp_returncode": int(print_result.returncode),
                "lp_stdout": str(print_result.stdout or "").strip(),
                "lp_stderr": str(print_result.stderr or "").strip(),
            }
        )

    meta_text = json.dumps(metadata, ensure_ascii=False, indent=2)
    meta_path.write_text(meta_text, encoding="utf-8")
    latest_meta_path.write_text(meta_text, encoding="utf-8")
    return metadata


def _send_raw_print_job(printer_name: str, raw_content: Any, copies: int = 1) -> None:
    with NamedTemporaryFile(prefix="bale-label-", suffix=".tspl", delete=False, mode="wb") as handle:
        temp_path = handle.name
        if isinstance(raw_content, bytes):
            payload = raw_content
        else:
            payload = str(raw_content or "").encode("ascii", errors="ignore")
        if max(int(copies or 1), 1) > 1:
            payload = payload * max(int(copies or 1), 1)
        handle.write(payload)
    try:
        try:
            print_result = subprocess.run(
                ["/usr/bin/lp", "-d", printer_name, "-o", "raw", temp_path],
                capture_output=True,
                text=True,
                check=False,
            )
        except Exception as exc:
            _build_print_debug_snapshot(
                printer_name,
                payload,
                copies=max(int(copies or 1), 1),
                error=str(exc),
            )
            raise
        _build_print_debug_snapshot(
            printer_name,
            payload,
            copies=max(int(copies or 1), 1),
            print_result=print_result,
        )
        if print_result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=(print_result.stderr or print_result.stdout or "TSPL 原始打印任务发送失败").strip(),
            )
    finally:
        try:
            os.remove(temp_path)
        except OSError:
            pass


def _extract_access_token(
    authorization: Optional[str] = None,
    access_token: Optional[str] = None,
) -> str:
    if access_token:
        return access_token
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")

    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Expected Bearer token")
    return token.strip()


def _require_current_user(
    authorization: Optional[str] = None,
    access_token: Optional[str] = None,
) -> dict[str, Any]:
    token = _extract_access_token(authorization=authorization, access_token=access_token)
    return state.get_authenticated_user(token)


def _is_store_clerk_user(user: dict[str, Any]) -> bool:
    return str((user or {}).get("role_code") or "").strip().lower() in {"store_clerk", "clerk", "store_staff", "sales_clerk"}


def _enforce_store_clerk_bale_access(current_user: dict[str, Any], bale: dict[str, Any]) -> None:
    if not _is_store_clerk_user(current_user):
        return
    assigned_employee = str(bale.get("assigned_employee") or "").strip().lower()
    username = str(current_user.get("username") or "").strip().lower()
    if not assigned_employee or assigned_employee != username:
        raise HTTPException(status_code=403, detail="当前 bale 没有分配给这个店员")


def _enforce_store_clerk_session_access(current_user: dict[str, Any], session: dict[str, Any]) -> None:
    if not _is_store_clerk_user(current_user):
        return
    assigned_employee = str(session.get("assigned_employee") or "").strip().lower()
    username = str(current_user.get("username") or "").strip().lower()
    if not assigned_employee or assigned_employee != username:
        raise HTTPException(status_code=403, detail="当前上架会话没有分配给这个店员")


def _normalize_import_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def _build_product_import_row(row_no: int, supplier_name: Any, category_main: Any, category_sub: Any) -> ProductImportRow:
    supplier = str(supplier_name or "").strip()
    main = str(category_main or "").strip()
    sub = str(category_sub or "").strip()
    issues: list[str] = []
    if not supplier:
        issues.append("缺少供应商")
    if not main:
        issues.append("缺少商品大类")
    if not sub:
        issues.append("缺少商品小类")
    return ProductImportRow(
        row_no=row_no,
        supplier_name=supplier,
        category_main=main,
        category_sub=sub,
        product_name=sub or main,
        valid=not issues,
        issues=issues,
    )


def _parse_product_import_rows(file_name: str, content: bytes) -> list[ProductImportRow]:
    normalized_name = str(file_name or "").strip().lower()
    rows: list[ProductImportRow] = []
    if normalized_name.endswith(".csv"):
        reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
        for index, row in enumerate(reader, start=2):
            source = { _normalize_import_header(key): value for key, value in (row or {}).items() }
            rows.append(
                _build_product_import_row(
                    index,
                    source.get("供应商") or source.get("supplier") or source.get("suppliername"),
                    source.get("商品大类") or source.get("大类") or source.get("categorymain"),
                    source.get("商品小类") or source.get("小类") or source.get("categorysub"),
                )
            )
        return rows

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    header_cells = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {
        _normalize_import_header(value): index
        for index, value in enumerate(header_cells)
        if str(value or "").strip()
    }
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(str(value or "").strip() for value in values):
            continue
        supplier_name = values[header_map.get("供应商", header_map.get("supplier", -1))] if header_map.get("供应商", header_map.get("supplier", -1)) >= 0 else ""
        category_main = values[header_map.get("商品大类", header_map.get("大类", header_map.get("categorymain", -1)))] if header_map.get("商品大类", header_map.get("大类", header_map.get("categorymain", -1))) >= 0 else ""
        category_sub = values[header_map.get("商品小类", header_map.get("小类", header_map.get("categorysub", -1)))] if header_map.get("商品小类", header_map.get("小类", header_map.get("categorysub", -1))) >= 0 else ""
        rows.append(_build_product_import_row(row_no, supplier_name, category_main, category_sub))
    return rows


def _build_china_source_import_row(
    row_no: int,
    supplier_name: Any,
    package_code: Any,
    category_main: Any,
    category_main_zh: Any,
    category_sub_zh: Any,
    category_sub: Any,
    package_count: Any,
    unit_weight_kg: Any,
    unit_cost_amount: Any,
    unit_cost_currency: Any,
) -> ChinaSourceImportRow:
    supplier = str(supplier_name or "").strip()
    package = str(package_code or "").strip()
    main = str(category_main or "").strip()
    main_zh = str(category_main_zh or "").strip()
    sub_zh = str(category_sub_zh or "").strip()
    sub = str(category_sub or "").strip()
    issues: list[str] = []
    try:
        package_count_value = int(float(package_count or 0))
    except (TypeError, ValueError):
        package_count_value = 0
    try:
        unit_weight_value = float(unit_weight_kg or 0)
    except (TypeError, ValueError):
        unit_weight_value = 0
    try:
        unit_cost_value = float(unit_cost_amount or 0)
    except (TypeError, ValueError):
        unit_cost_value = 0
    currency = str(unit_cost_currency or "CNY").strip().upper() or "CNY"
    if not supplier:
        issues.append("缺少供应商")
    if not package:
        issues.append("缺少包裹编码")
    if not main:
        issues.append("缺少商品大类")
    if not sub:
        issues.append("缺少商品小类")
    if package_count_value <= 0:
        issues.append("缺少包数")
    if unit_weight_value <= 0:
        issues.append("缺少单包重量")
    if unit_cost_value < 0:
        issues.append("包裹单价不能为负数")
    return ChinaSourceImportRow(
        row_no=row_no,
        supplier_name=supplier,
        package_code=package,
        category_main=main,
        category_main_zh=main_zh,
        category_sub_zh=sub_zh,
        category_sub=sub,
        package_count=package_count_value,
        unit_weight_kg=unit_weight_value,
        unit_cost_amount=unit_cost_value,
        unit_cost_currency=currency,
        valid=not issues,
        issues=issues,
    )


def _is_china_source_instruction_row(
    supplier_name: Any,
    category_main: Any,
    category_sub: Any,
    package_count: Any,
    unit_weight_kg: Any,
    unit_cost_amount: Any,
) -> bool:
    supplier = str(supplier_name or "").strip()
    main = str(category_main or "").strip()
    sub = str(category_sub or "").strip()
    text_blob = " ".join(part for part in [supplier, main, sub] if part).strip()
    if not text_blob:
        return False
    try:
        package_count_value = float(package_count or 0)
    except (TypeError, ValueError):
        package_count_value = 0
    try:
        unit_weight_value = float(unit_weight_kg or 0)
    except (TypeError, ValueError):
        unit_weight_value = 0
    try:
        unit_cost_value = float(unit_cost_amount or 0)
    except (TypeError, ValueError):
        unit_cost_value = 0
    if package_count_value > 0 or unit_weight_value > 0 or unit_cost_value > 0:
        return False
    instruction_patterns = [
        r"^说明[:：]?",
        r"^\d+\s*[.、]",
        r"模板",
        r"上传后",
        r"只导入",
        r"货币支持",
        r"继续修改",
        r"填写",
    ]
    return any(re.search(pattern, text_blob, flags=re.IGNORECASE) for pattern in instruction_patterns)


def _parse_china_source_import_rows(file_name: str, content: bytes) -> list[ChinaSourceImportRow]:
    normalized_name = str(file_name or "").strip().lower()
    rows: list[ChinaSourceImportRow] = []
    if normalized_name.endswith(".csv"):
        reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
        for index, row in enumerate(reader, start=2):
            source = {_normalize_import_header(key): value for key, value in (row or {}).items()}
            supplier_name = source.get("供应商") or source.get("supplier") or source.get("suppliername")
            package_code = source.get("包裹编码") or source.get("packagecode")
            category_main = source.get("大类英") or source.get("商品大类") or source.get("大类") or source.get("categorymain")
            category_main_zh = source.get("大类中") or source.get("categorymainzh")
            category_sub_zh = source.get("小类中") or source.get("categorysubzh")
            category_sub = source.get("小类英") or source.get("商品小类") or source.get("小类") or source.get("categorysub")
            package_count = source.get("包数") or source.get("packagecount")
            unit_weight_kg = source.get("单包重量kg") or source.get("单包重量") or source.get("unitweightkg")
            unit_cost_amount = source.get("包裹单价") or source.get("unitcostamount") or source.get("金额")
            if _is_china_source_instruction_row(
                supplier_name,
                category_main,
                category_sub,
                package_count,
                unit_weight_kg,
                unit_cost_amount,
            ):
                continue
            rows.append(
                _build_china_source_import_row(
                    index,
                    supplier_name,
                    package_code,
                    category_main,
                    category_main_zh,
                    category_sub_zh,
                    category_sub,
                    package_count,
                    unit_weight_kg,
                    unit_cost_amount,
                    source.get("货币") or source.get("unitcostcurrency") or source.get("currency"),
                )
            )
        return rows

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    header_cells = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {
        _normalize_import_header(value): index
        for index, value in enumerate(header_cells)
        if str(value or "").strip()
    }

    def cell(values: tuple[Any, ...], *keys: str) -> Any:
        for key in keys:
            idx = header_map.get(key, -1)
            if idx >= 0 and idx < len(values):
                return values[idx]
        return ""

    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(str(value or "").strip() for value in values):
            continue
        supplier_name = cell(values, "供应商", "supplier", "suppliername")
        package_code = cell(values, "包裹编码", "packagecode")
        category_main = cell(values, "大类英", "商品大类", "大类", "categorymain")
        category_main_zh = cell(values, "大类中", "categorymainzh")
        category_sub_zh = cell(values, "小类中", "categorysubzh")
        category_sub = cell(values, "小类英", "商品小类", "小类", "categorysub")
        package_count = cell(values, "包数", "packagecount")
        unit_weight_kg = cell(values, "单包重量kg", "单包重量", "unitweightkg")
        unit_cost_amount = cell(values, "包裹单价", "unitcostamount", "金额")
        if _is_china_source_instruction_row(
            supplier_name,
            category_main,
            category_sub,
            package_count,
            unit_weight_kg,
            unit_cost_amount,
        ):
            continue
        rows.append(
            _build_china_source_import_row(
                row_no,
                supplier_name,
                package_code,
                category_main,
                category_main_zh,
                category_sub_zh,
                category_sub,
                package_count,
                unit_weight_kg,
                unit_cost_amount,
                cell(values, "货币", "unitcostcurrency", "currency"),
            )
        )
    return rows


def _build_bale_sales_pricing_workbook(rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待售包裹成本表"
    sheet["A1"] = "待售 bale 成本与毛利编辑表"
    sheet["A2"] = "说明：source_cost_kes 来自船单录入与后续成本池；editable_cost_kes / downstream_cost_kes / margin_rate 可用于销售编辑。"
    headers = [
        "entry_id",
        "source_type",
        "bale_barcode",
        "shipment_no",
        "supplier_name",
        "category_main",
        "category_sub",
        "weight_kg",
        "source_cost_kes",
        "editable_cost_kes",
        "downstream_cost_kes",
        "total_cost_kes",
        "margin_rate",
        "target_sale_price_kes",
        "status",
        "pricing_note",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column_index, value=header)
    for row_index, row in enumerate(rows, start=5):
        values = [
            row.get("entry_id", ""),
            row.get("source_type", ""),
            row.get("bale_barcode", ""),
            row.get("shipment_no", ""),
            row.get("supplier_name", ""),
            row.get("category_main", ""),
            row.get("category_sub", ""),
            row.get("weight_kg", 0),
            row.get("source_cost_kes", 0),
            row.get("editable_cost_kes", 0),
            row.get("downstream_cost_kes", 0),
            row.get("total_cost_kes", 0),
            row.get("margin_rate", 0),
            row.get("target_sale_price_kes", 0),
            row.get("status", ""),
            row.get("pricing_note", ""),
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_bale_sales_order_workbook(order: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bales销售单"
    sheet["A1"] = "Bales 销售出库单"
    sheet["A2"] = "order_no"
    sheet["B2"] = order.get("order_no", "")
    sheet["A3"] = "sold_by"
    sheet["B3"] = order.get("sold_by", "")
    sheet["A4"] = "customer_name"
    sheet["B4"] = order.get("customer_name", "")
    sheet["A5"] = "customer_contact"
    sheet["B5"] = order.get("customer_contact", "")
    sheet["A6"] = "payment_method"
    sheet["B6"] = order.get("payment_method", "")
    headers = [
        "bale_barcode",
        "shipment_no",
        "supplier_name",
        "total_cost_kes",
        "sale_price_kes",
        "profit_kes",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=8, column=column_index, value=header)
    for row_index, item in enumerate(order.get("items", []) or [], start=9):
        values = [
            item.get("bale_barcode", ""),
            item.get("shipment_no", ""),
            item.get("supplier_name", ""),
            item.get("total_cost_kes", 0),
            item.get("sale_price_kes", 0),
            item.get("profit_kes", 0),
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    summary_row = 10 + len(order.get("items", []) or [])
    sheet.cell(row=summary_row, column=1, value="total_amount_kes")
    sheet.cell(row=summary_row, column=2, value=order.get("total_amount_kes", 0))
    sheet.cell(row=summary_row + 1, column=1, value="total_profit_kes")
    sheet.cell(row=summary_row + 1, column=2, value=order.get("total_profit_kes", 0))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.get("/health", response_model=MessageResponse, tags=["system"])
def health_check() -> MessageResponse:
    return MessageResponse(message=f"{settings.app_name} is running")


@router.post("/system/reset-test-history", response_model=MessageResponse, tags=["system"])
def reset_test_history(
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    result = state.reset_test_history(current_user["username"])
    return MessageResponse(message=result["message"])


@router.post("/system/generate-warehouse-mainflow-demo", response_model=WarehouseMainflowDemoResponse, tags=["system", "warehouse"])
def generate_warehouse_mainflow_demo(
    authorization: Optional[str] = Header(default=None),
) -> WarehouseMainflowDemoResponse:
    current_user = _require_current_user(authorization=authorization)
    return WarehouseMainflowDemoResponse(**state.generate_warehouse_mainflow_demo(current_user["username"]))


@router.post("/admin/tools/raw-bale-machine-code-repair", tags=["system", "admin"])
def repair_raw_bale_machine_codes(
    payload: dict[str, Any],
    authorization: Optional[str] = Header(default=None),
) -> dict[str, Any]:
    current_user = _require_current_user(authorization=authorization)
    return state.repair_raw_bale_machine_codes(
        dry_run=bool(payload.get("dry_run", True)),
        actor_username=current_user["username"],
    )


@router.post(
    "/system/generate-store-replenishment-demo",
    response_model=StoreReplenishmentDemoResponse,
    tags=["system", "warehouse", "transfers", "sales"],
)
def generate_store_replenishment_demo(
    authorization: Optional[str] = Header(default=None),
) -> StoreReplenishmentDemoResponse:
    current_user = _require_current_user(authorization=authorization)
    return StoreReplenishmentDemoResponse(**state.generate_store_replenishment_demo(current_user["username"]))


@router.post("/auth/login", response_model=LoginResponse, tags=["auth"])
def login(payload: LoginRequest) -> LoginResponse:
    session = state.authenticate_user(payload.username, payload.password)
    return LoginResponse(**session)


@router.get("/auth/me", response_model=SessionUserResponse, tags=["auth"])
def get_session_user(
    authorization: Optional[str] = Header(default=None),
) -> SessionUserResponse:
    current_user = _require_current_user(authorization=authorization)
    return SessionUserResponse(**current_user)


@router.post("/auth/logout", response_model=MessageResponse, tags=["auth"])
def logout(
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    token = _extract_access_token(authorization=authorization)
    current_user = state.get_authenticated_user(token)
    state.logout_user(token)
    return MessageResponse(message=f"User {current_user['username']} logged out")


@router.get("/dashboard/summary", response_model=list[SummaryCard], tags=["dashboard"])
def get_dashboard_summary(
    authorization: Optional[str] = Header(default=None),
) -> list[SummaryCard]:
    _require_current_user(authorization=authorization)
    return [SummaryCard(**row) for row in state.get_dashboard_summary()]


@router.get(
    "/dashboard/store-operating-summary",
    response_model=list[StoreOperatingSummaryResponse],
    tags=["dashboard"],
)
def get_store_operating_summary(
    authorization: Optional[str] = Header(default=None),
) -> list[StoreOperatingSummaryResponse]:
    _require_current_user(authorization=authorization)
    return [StoreOperatingSummaryResponse(**row) for row in state.get_store_operating_summary()]


@router.get(
    "/dashboard/store-closing-checklist/{store_code}",
    response_model=StoreClosingChecklistResponse,
    tags=["dashboard", "pos"],
)
def get_store_closing_checklist(
    store_code: str,
    authorization: Optional[str] = Header(default=None),
) -> StoreClosingChecklistResponse:
    _require_current_user(authorization=authorization)
    return StoreClosingChecklistResponse(**state.get_store_closing_checklist(store_code))


@router.get("/stores", response_model=list[StoreResponse], tags=["stores"])
def list_stores(
    authorization: Optional[str] = Header(default=None),
) -> list[StoreResponse]:
    _require_current_user(authorization=authorization)
    return [StoreResponse(**store) for store in state.list_stores()]


@router.get("/suppliers", response_model=list[SupplierResponse], tags=["suppliers"])
def list_suppliers(
    authorization: Optional[str] = Header(default=None),
) -> list[SupplierResponse]:
    _require_current_user(authorization=authorization)
    return [SupplierResponse(**supplier) for supplier in state.list_suppliers()]


@router.post("/suppliers", response_model=SupplierResponse, tags=["suppliers"])
def create_supplier(
    payload: SupplierCreate,
    authorization: Optional[str] = Header(default=None),
) -> SupplierResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return SupplierResponse(**state.create_supplier(data))


@router.get("/cargo-types", response_model=list[CargoTypeResponse], tags=["cargo_types"])
def list_cargo_types(
    authorization: Optional[str] = Header(default=None),
) -> list[CargoTypeResponse]:
    _require_current_user(authorization=authorization)
    return [CargoTypeResponse(**row) for row in state.list_cargo_types()]


@router.post("/cargo-types", response_model=CargoTypeResponse, tags=["cargo_types"])
def create_cargo_type(
    payload: CargoTypeCreate,
    authorization: Optional[str] = Header(default=None),
) -> CargoTypeResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return CargoTypeResponse(**state.create_cargo_type(data))


@router.post("/stores", response_model=StoreResponse, tags=["stores"])
def create_store(
    payload: StoreCreate,
    authorization: Optional[str] = Header(default=None),
) -> StoreResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return StoreResponse(**state.create_store(data))


@router.post("/stores/site-recommendation", response_model=StoreSiteRecommendationResponse, tags=["stores"])
def recommend_store_site(
    payload: StoreSiteRecommendationRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreSiteRecommendationResponse:
    _require_current_user(authorization=authorization)
    return StoreSiteRecommendationResponse(**state.recommend_store_site(payload.model_dump()))


@router.get("/roles", response_model=list[RoleResponse], tags=["users"])
def list_roles(
    authorization: Optional[str] = Header(default=None),
) -> list[RoleResponse]:
    _require_current_user(authorization=authorization)
    return [RoleResponse(**role) for role in ROLE_DEFINITIONS]


@router.post("/users", response_model=UserResponse, tags=["users"])
def create_user(
    payload: UserCreate,
    authorization: Optional[str] = Header(default=None),
) -> UserResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return UserResponse(**state.create_user(data))


@router.get("/users", response_model=list[UserResponse], tags=["users"])
def list_users(
    authorization: Optional[str] = Header(default=None),
) -> list[UserResponse]:
    _require_current_user(authorization=authorization)
    return [UserResponse(**user) for user in state.list_users()]


@router.patch("/users/{user_id}", response_model=UserResponse, tags=["users"])
def update_user(
    user_id: int,
    payload: UserUpdate,
    authorization: Optional[str] = Header(default=None),
) -> UserResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump(exclude_unset=True)
    data["updated_by"] = current_user["username"]
    return UserResponse(**state.update_user(user_id, data))


@router.delete("/users/{user_id}", response_model=UserResponse, tags=["users"])
def deactivate_user(
    user_id: int,
    authorization: Optional[str] = Header(default=None),
) -> UserResponse:
    current_user = _require_current_user(authorization=authorization)
    return UserResponse(**state.deactivate_user(user_id, current_user["username"]))


@router.get(
    "/stores/rack-template",
    response_model=list[StoreRackTemplateResponse],
    tags=["stores"],
)
def get_store_rack_template(
    authorization: Optional[str] = Header(default=None),
) -> list[StoreRackTemplateResponse]:
    _require_current_user(authorization=authorization)
    return [StoreRackTemplateResponse(**row) for row in STORE_RACK_TEMPLATE]


@router.get(
    "/settings/barcode",
    response_model=BarcodeSettingsResponse,
    tags=["settings"],
)
def get_barcode_settings(
    authorization: Optional[str] = Header(default=None),
) -> BarcodeSettingsResponse:
    _require_current_user(authorization=authorization)
    return BarcodeSettingsResponse(**BARCODE_SETTINGS)


@router.get(
    "/settings/label-templates",
    response_model=list[LabelTemplateResponse],
    tags=["settings"],
)
def list_label_templates(
    authorization: Optional[str] = Header(default=None),
) -> list[LabelTemplateResponse]:
    _require_current_user(authorization=authorization)
    return [LabelTemplateResponse(**template) for template in state.list_label_templates()]


@router.post(
    "/settings/label-templates",
    response_model=LabelTemplateResponse,
    tags=["settings"],
)
def create_label_template(
    payload: LabelTemplateSaveRequest,
    authorization: Optional[str] = Header(default=None),
) -> LabelTemplateResponse:
    current_user = _require_current_user(authorization=authorization)
    result = state.save_label_template(payload.model_dump(), updated_by=current_user["username"])
    return LabelTemplateResponse(**result)


@router.put(
    "/settings/label-templates/{template_code}",
    response_model=LabelTemplateResponse,
    tags=["settings"],
)
def update_label_template(
    template_code: str,
    payload: LabelTemplateSaveRequest,
    authorization: Optional[str] = Header(default=None),
) -> LabelTemplateResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["template_code"] = template_code
    result = state.save_label_template(data, updated_by=current_user["username"])
    return LabelTemplateResponse(**result)


@router.post(
    "/pricing/rules",
    response_model=PriceRuleResponse,
    tags=["pricing"],
)
def create_price_rule(
    payload: PriceRuleCreate,
    authorization: Optional[str] = Header(default=None),
) -> PriceRuleResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return PriceRuleResponse(**state.create_price_rule(data))


@router.get(
    "/pricing/rules",
    response_model=list[PriceRuleResponse],
    tags=["pricing"],
)
def list_price_rules(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[PriceRuleResponse]:
    _require_current_user(authorization=authorization)
    return [
        PriceRuleResponse(**row)
        for row in state.list_price_rules(store_code=store_code, status=status)
    ]


@router.post("/products", response_model=ProductResponse, tags=["products"])
def create_product(
    payload: ProductCreate,
    authorization: Optional[str] = Header(default=None),
) -> ProductResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return ProductResponse(**state.create_product(data))


@router.post("/products/import-preview", response_model=ProductImportPreviewResponse, tags=["products"])
async def preview_product_import(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
) -> ProductImportPreviewResponse:
    _require_current_user(authorization=authorization)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空。")
    try:
        rows = _parse_product_import_rows(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法识别上传文件，请使用模板重新导出后再试。{exc}") from exc
    valid_rows = sum(1 for row in rows if row.valid)
    return ProductImportPreviewResponse(
        file_name=file.filename or "",
        total_rows=len(rows),
        valid_rows=valid_rows,
        invalid_rows=len(rows) - valid_rows,
        rows=rows,
    )


@router.post("/products/bulk-create", response_model=list[ProductResponse], tags=["products"])
def bulk_create_products(
    payload: ProductBulkCreateRequest,
    authorization: Optional[str] = Header(default=None),
) -> list[ProductResponse]:
    current_user = _require_current_user(authorization=authorization)
    created: list[ProductResponse] = []
    for item in payload.items:
        data = item.model_dump()
        data["created_by"] = current_user["username"]
        created.append(ProductResponse(**state.create_product(data)))
    return created


@router.post("/china-source/import-preview", response_model=ChinaSourceImportPreviewResponse, tags=["warehouse"])
async def preview_china_source_import(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
) -> ChinaSourceImportPreviewResponse:
    _require_current_user(authorization=authorization)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空。")
    try:
        rows = _parse_china_source_import_rows(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法识别上传文件，请使用模板重新导出后再试。{exc}") from exc
    valid_rows = sum(1 for row in rows if row.valid)
    return ChinaSourceImportPreviewResponse(
        file_name=file.filename or "",
        total_rows=len(rows),
        valid_rows=valid_rows,
        invalid_rows=len(rows) - valid_rows,
        rows=rows,
    )


@router.get("/warehouse/china-sources", response_model=list[ChinaSourceRecordResponse], tags=["warehouse"])
def list_china_source_records(
    authorization: Optional[str] = Header(default=None),
) -> list[ChinaSourceRecordResponse]:
    _require_current_user(authorization=authorization)
    return [ChinaSourceRecordResponse(**row) for row in state.list_china_source_records()]


@router.post("/warehouse/china-sources", response_model=ChinaSourceRecordResponse, tags=["warehouse"])
def create_or_update_china_source_record(
    payload: ChinaSourceRecordCreate,
    authorization: Optional[str] = Header(default=None),
) -> ChinaSourceRecordResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return ChinaSourceRecordResponse(**state.create_or_update_china_source_record(data, created_by=current_user["username"]))


@router.put(
    "/warehouse/china-sources/{source_pool_token}/cost",
    response_model=ChinaSourceRecordResponse,
    tags=["warehouse"],
)
def update_china_source_cost(
    source_pool_token: str,
    payload: ChinaSourceCostUpdateRequest,
    authorization: Optional[str] = Header(default=None),
) -> ChinaSourceRecordResponse:
    current_user = _require_current_user(authorization=authorization)
    return ChinaSourceRecordResponse(
        **state.update_china_source_cost(
            source_pool_token,
            payload.model_dump(),
            updated_by=current_user["username"],
        )
    )


@router.get("/warehouse/apparel-piece-weights", response_model=list[ApparelPieceWeightResponse], tags=["warehouse"])
def list_apparel_piece_weights(
    authorization: Optional[str] = Header(default=None),
) -> list[ApparelPieceWeightResponse]:
    _require_current_user(authorization=authorization)
    return [ApparelPieceWeightResponse(**row) for row in state.list_apparel_piece_weights()]


@router.post("/warehouse/apparel-piece-weights", response_model=ApparelPieceWeightResponse, tags=["warehouse"])
def upsert_apparel_piece_weight(
    payload: ApparelPieceWeightCreate,
    authorization: Optional[str] = Header(default=None),
) -> ApparelPieceWeightResponse:
    current_user = _require_current_user(authorization=authorization)
    return ApparelPieceWeightResponse(
        **state.upsert_apparel_piece_weight(payload.model_dump(), updated_by=current_user["username"])
    )


@router.delete(
    "/warehouse/apparel-piece-weights/{category_main}/{category_sub}",
    response_model=MessageResponse,
    tags=["warehouse"],
)
def delete_apparel_piece_weight(
    category_main: str,
    category_sub: str,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    state.delete_apparel_piece_weight(category_main, category_sub, deleted_by=current_user["username"])
    return MessageResponse(message=f"Deleted apparel piece weight for {category_main} / {category_sub}")


@router.get("/warehouse/apparel-default-costs", response_model=list[ApparelDefaultCostResponse], tags=["warehouse"])
def list_apparel_default_costs(
    authorization: Optional[str] = Header(default=None),
) -> list[ApparelDefaultCostResponse]:
    _require_current_user(authorization=authorization)
    return [ApparelDefaultCostResponse(**row) for row in state.list_apparel_default_costs()]


@router.post("/warehouse/apparel-default-costs", response_model=ApparelDefaultCostResponse, tags=["warehouse"])
def upsert_apparel_default_cost(
    payload: ApparelDefaultCostCreate,
    authorization: Optional[str] = Header(default=None),
) -> ApparelDefaultCostResponse:
    current_user = _require_current_user(authorization=authorization)
    return ApparelDefaultCostResponse(
        **state.upsert_apparel_default_cost(payload.model_dump(), updated_by=current_user["username"])
    )


@router.delete(
    "/warehouse/apparel-default-costs/{category_main}/{category_sub}/{grade}",
    response_model=MessageResponse,
    tags=["warehouse"],
)
def delete_apparel_default_cost(
    category_main: str,
    category_sub: str,
    grade: str,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    state.delete_apparel_default_cost(category_main, category_sub, grade, deleted_by=current_user["username"])
    return MessageResponse(message=f"Deleted apparel default cost for {category_main} / {category_sub} / {grade}")


@router.get("/warehouse/apparel-sorting-racks", response_model=list[ApparelSortingRackResponse], tags=["warehouse"])
def list_apparel_sorting_racks(
    authorization: Optional[str] = Header(default=None),
) -> list[ApparelSortingRackResponse]:
    _require_current_user(authorization=authorization)
    return [ApparelSortingRackResponse(**row) for row in state.list_apparel_sorting_racks()]


@router.post("/warehouse/apparel-sorting-racks", response_model=ApparelSortingRackResponse, tags=["warehouse"])
def upsert_apparel_sorting_rack(
    payload: ApparelSortingRackCreate,
    authorization: Optional[str] = Header(default=None),
) -> ApparelSortingRackResponse:
    current_user = _require_current_user(authorization=authorization)
    return ApparelSortingRackResponse(
        **state.upsert_apparel_sorting_rack(payload.model_dump(), updated_by=current_user["username"])
    )


@router.delete(
    "/warehouse/apparel-sorting-racks/{category_main}/{category_sub}/{grade}/{default_cost_kes}",
    response_model=MessageResponse,
    tags=["warehouse"],
)
def delete_apparel_sorting_rack(
    category_main: str,
    category_sub: str,
    grade: str,
    default_cost_kes: float,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    state.delete_apparel_sorting_rack(
        category_main,
        category_sub,
        grade,
        default_cost_kes,
        deleted_by=current_user["username"],
    )
    return MessageResponse(
        message=f"Deleted apparel sorting rack for {category_main} / {category_sub} / {grade} / {default_cost_kes:.2f}"
    )


@router.post("/warehouse/parcel-batches", response_model=ParcelBatchResponse, tags=["warehouse"])
def create_parcel_batch(
    payload: ParcelBatchCreate,
    authorization: Optional[str] = Header(default=None),
) -> ParcelBatchResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["received_by"] = current_user["username"]
    return ParcelBatchResponse(**state.create_parcel_batch(data))


@router.post("/warehouse/inbound-shipments", response_model=InboundShipmentResponse, tags=["warehouse"])
def create_inbound_shipment(
    payload: InboundShipmentCreate,
    authorization: Optional[str] = Header(default=None),
) -> InboundShipmentResponse:
    _require_current_user(authorization=authorization)
    return InboundShipmentResponse(**state.create_inbound_shipment(payload.model_dump()))


@router.get("/warehouse/inbound-shipments", response_model=list[InboundShipmentResponse], tags=["warehouse"])
def list_inbound_shipments(
    shipment_type: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[InboundShipmentResponse]:
    _require_current_user(authorization=authorization)
    return [InboundShipmentResponse(**row) for row in state.list_inbound_shipments(shipment_type=shipment_type)]


@router.get("/warehouse/parcel-batches", response_model=list[ParcelBatchResponse], tags=["warehouse"])
def list_parcel_batches(
    status: Optional[str] = Query(default=None),
    shipment_no: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[ParcelBatchResponse]:
    _require_current_user(authorization=authorization)
    return [
        ParcelBatchResponse(**row)
        for row in state.list_parcel_batches(status=status, shipment_no=shipment_no)
    ]


@router.post(
    "/warehouse/inbound-shipments/{shipment_no}/confirm-intake",
    response_model=InboundShipmentResponse,
    tags=["warehouse"],
)
def confirm_inbound_shipment_intake(
    shipment_no: str,
    payload: InboundShipmentIntakeConfirmRequest,
    authorization: Optional[str] = Header(default=None),
) -> InboundShipmentResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["confirmed_by"] = current_user["username"]
    return InboundShipmentResponse(**state.confirm_inbound_shipment_intake(shipment_no, data))


@router.post(
    "/warehouse/inbound-shipments/{shipment_no}/generate-bale-barcodes",
    response_model=list[BaleBarcodeResponse],
    tags=["warehouse"],
)
def generate_bale_barcodes(
    shipment_no: str,
    authorization: Optional[str] = Header(default=None),
) -> list[BaleBarcodeResponse]:
    current_user = _require_current_user(authorization=authorization)
    return [BaleBarcodeResponse(**row) for row in state.generate_bale_barcodes(shipment_no, current_user["username"])]


@router.get(
    "/warehouse/bale-barcodes",
    response_model=list[BaleBarcodeResponse],
    tags=["warehouse"],
)
def list_bale_barcodes(
    shipment_no: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[BaleBarcodeResponse]:
    _require_current_user(authorization=authorization)
    return [
        BaleBarcodeResponse(**row)
        for row in state.list_bale_barcodes(shipment_no=shipment_no, status=status)
    ]


@router.get("/warehouse/raw-bales", response_model=list[RawBaleStockResponse], tags=["warehouse"])
def list_raw_bales(
    shipment_no: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    destination_judgement: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[RawBaleStockResponse]:
    _require_current_user(authorization=authorization)
    return [
        RawBaleStockResponse(**row)
        for row in state.list_raw_bales(
            shipment_no=shipment_no,
            status=status,
            destination_judgement=destination_judgement,
        )
    ]


@router.post("/warehouse/raw-bales/{bale_barcode}/route-to-sorting", response_model=RawBaleStockResponse, tags=["warehouse"])
def route_raw_bale_to_sorting(
    bale_barcode: str,
    payload: RawBaleRouteRequest,
    authorization: Optional[str] = Header(default=None),
) -> RawBaleStockResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return RawBaleStockResponse(**state.route_raw_bale_to_sorting(bale_barcode, data))


@router.post(
    "/warehouse/raw-bales/{bale_barcode}/route-to-bale-sales-pool",
    response_model=RawBaleStockResponse,
    tags=["warehouse"],
)
def route_raw_bale_to_bale_sales_pool(
    bale_barcode: str,
    payload: RawBaleRouteRequest,
    authorization: Optional[str] = Header(default=None),
) -> RawBaleStockResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return RawBaleStockResponse(**state.route_raw_bale_to_bale_sales_pool(bale_barcode, data))


@router.get("/bale-sales/candidates", response_model=list[BaleSalesCandidateResponse], tags=["bale-sales"])
def list_bale_sales_candidates(
    shipment_no: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    source_type: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[BaleSalesCandidateResponse]:
    _require_current_user(authorization=authorization)
    return [
        BaleSalesCandidateResponse(**row)
        for row in state.list_bale_sales_candidates(
            shipment_no=shipment_no,
            status=status,
            source_type=source_type,
        )
    ]


@router.post(
    "/bale-sales/candidates/{entry_id}/pricing",
    response_model=BaleSalesCandidateResponse,
    tags=["bale-sales"],
)
def update_bale_sales_candidate_pricing(
    entry_id: str,
    payload: BaleSalesCandidatePricingUpdateRequest,
    authorization: Optional[str] = Header(default=None),
) -> BaleSalesCandidateResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return BaleSalesCandidateResponse(**state.update_bale_sales_candidate_pricing(entry_id, data))


@router.get("/bale-sales/exports/pricing-sheet.xlsx", tags=["bale-sales"])
def export_bale_sales_pricing_sheet(
    shipment_no: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    source_type: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> StreamingResponse:
    _require_current_user(authorization=authorization)
    rows = state.list_bale_sales_candidates(
        shipment_no=shipment_no,
        status=status,
        source_type=source_type,
    )
    content = _build_bale_sales_pricing_workbook(rows)
    file_name = f"bale-sales-pricing-{datetime.now(NAIROBI_TZ).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.post("/bale-sales/orders", response_model=BaleSalesOrderResponse, tags=["bale-sales"])
def create_bale_sales_order(
    payload: BaleSalesOrderCreate,
    authorization: Optional[str] = Header(default=None),
) -> BaleSalesOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return BaleSalesOrderResponse(**state.create_bale_sales_order(data))


@router.get("/bale-sales/orders", response_model=list[BaleSalesOrderResponse], tags=["bale-sales"])
def list_bale_sales_orders(
    authorization: Optional[str] = Header(default=None),
) -> list[BaleSalesOrderResponse]:
    _require_current_user(authorization=authorization)
    return [BaleSalesOrderResponse(**row) for row in state.list_bale_sales_orders()]


@router.get("/bale-sales/orders/{order_no}", response_model=BaleSalesOrderResponse, tags=["bale-sales"])
def get_bale_sales_order(
    order_no: str,
    authorization: Optional[str] = Header(default=None),
) -> BaleSalesOrderResponse:
    _require_current_user(authorization=authorization)
    return BaleSalesOrderResponse(**state.get_bale_sales_order(order_no))


@router.get("/bale-sales/orders/{order_no}/sales-sheet.xlsx", tags=["bale-sales"])
def export_bale_sales_order_sheet(
    order_no: str,
    authorization: Optional[str] = Header(default=None),
) -> StreamingResponse:
    _require_current_user(authorization=authorization)
    order = state.get_bale_sales_order(order_no)
    content = _build_bale_sales_order_workbook(order)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{order_no}-sales-sheet.xlsx"'},
    )


@router.post(
    "/warehouse/bale-barcodes/print-jobs",
    response_model=BaleBarcodePrintResponse,
    tags=["warehouse", "printing"],
)
def queue_bale_barcode_print_jobs(
    payload: BaleBarcodePrintRequest,
    authorization: Optional[str] = Header(default=None),
) -> BaleBarcodePrintResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    result = state.queue_bale_barcode_print_jobs(
        shipment_no=data["shipment_no"],
        items=data["items"],
        requested_by=data["requested_by"],
        printer_name=data["printer_name"],
        template_code=data["template_code"],
    )
    return BaleBarcodePrintResponse(
        shipment_no=result["shipment_no"],
        print_jobs=[PrintJobResponse(**row) for row in result["print_jobs"]],
        total_selected_bales=result["total_selected_bales"],
        total_print_copies=result["total_print_copies"],
    )


@router.post(
    "/warehouse/store-prep-bales/{bale_no}/print-jobs",
    response_model=PrintJobResponse,
    tags=["warehouse", "printing"],
)
def queue_store_prep_bale_print_job(
    bale_no: str,
    payload: StorePrepBalePrintJobCreate,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return PrintJobResponse(
        **state.queue_store_prep_bale_print_job(
            bale_no=bale_no,
            requested_by=data["requested_by"],
            printer_name=data["printer_name"],
            template_code=data["template_code"],
            copies=data["copies"],
        )
    )


@router.post("/warehouse/sorting-tasks", response_model=SortingTaskResponse, tags=["warehouse"])
def create_sorting_task(
    payload: SortingTaskCreate,
    authorization: Optional[str] = Header(default=None),
) -> SortingTaskResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return SortingTaskResponse(**state.create_sorting_task(data))


@router.get("/warehouse/sorting-tasks", response_model=list[SortingTaskResponse], tags=["warehouse"])
def list_sorting_tasks(
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[SortingTaskResponse]:
    _require_current_user(authorization=authorization)
    return [SortingTaskResponse(**row) for row in state.list_sorting_tasks(status=status)]


@router.post("/warehouse/store-prep-bale-tasks", response_model=StorePrepBaleTaskResponse, tags=["warehouse"])
def create_store_prep_bale_task(
    payload: StorePrepBaleTaskCreate,
    authorization: Optional[str] = Header(default=None),
) -> StorePrepBaleTaskResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return StorePrepBaleTaskResponse(**state.create_store_prep_bale_task(data))


@router.get("/warehouse/store-prep-bale-tasks", response_model=list[StorePrepBaleTaskResponse], tags=["warehouse"])
def list_store_prep_bale_tasks(
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[StorePrepBaleTaskResponse]:
    _require_current_user(authorization=authorization)
    return [StorePrepBaleTaskResponse(**row) for row in state.list_store_prep_bale_tasks(status=status)]


@router.post("/warehouse/store-prep-bale-tasks/{task_no}/complete", response_model=StorePrepBaleTaskResponse, tags=["warehouse"])
def complete_store_prep_bale_task(
    task_no: str,
    payload: StorePrepBaleTaskCompleteRequest,
    authorization: Optional[str] = Header(default=None),
) -> StorePrepBaleTaskResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return StorePrepBaleTaskResponse(**state.complete_store_prep_bale_task(task_no, data))


@router.get("/warehouse/store-prep-bales", response_model=list[StorePrepBaleResponse], tags=["warehouse"])
def list_store_prep_bales(
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[StorePrepBaleResponse]:
    _require_current_user(authorization=authorization)
    return [StorePrepBaleResponse(**row) for row in state.list_store_prep_bales(status=status)]


@router.get("/warehouse/inventory-summary", response_model=WarehouseInventorySummaryResponse, tags=["warehouse"])
def get_warehouse_inventory_summary(
    authorization: Optional[str] = Header(default=None),
) -> WarehouseInventorySummaryResponse:
    _require_current_user(authorization=authorization)
    return WarehouseInventorySummaryResponse(**state.get_warehouse_inventory_summary())


@router.get("/warehouse/item-barcode-tokens", response_model=list[ItemBarcodeTokenResponse], tags=["warehouse"])
def list_item_barcode_tokens(
    status: Optional[str] = Query(default=None),
    task_no: Optional[str] = Query(default=None),
    shipment_no: Optional[str] = Query(default=None),
    store_dispatch_bale_no: Optional[str] = Query(default=None),
    store_code: Optional[str] = Query(default=None),
    assigned_employee: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[ItemBarcodeTokenResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_item_barcode_tokens(
        status=status,
        task_no=task_no,
        shipment_no=shipment_no,
        store_dispatch_bale_no=store_dispatch_bale_no,
        store_code=store_code,
        assigned_employee=assigned_employee,
    )
    return [ItemBarcodeTokenResponse(**row) for row in rows]


@router.get(
    "/warehouse/item-identity-ledger/{identity_no}",
    response_model=ItemIdentityLedgerResponse,
    tags=["warehouse"],
)
def get_item_identity_ledger(
    identity_no: str,
    authorization: Optional[str] = Header(default=None),
) -> ItemIdentityLedgerResponse:
    _require_current_user(authorization=authorization)
    return ItemIdentityLedgerResponse(**state.get_item_identity_ledger(identity_no))


@router.post(
    "/stores/{store_code}/retail-demo-seed",
    response_model=StoreRetailSeedResponse,
    tags=["stores", "sales"],
)
def seed_store_retail_demo(
    store_code: str,
    payload: StoreRetailSeedRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreRetailSeedResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["store_code"] = store_code
    data["seeded_by"] = current_user["username"]
    return StoreRetailSeedResponse(**state.seed_store_retail_samples(data))


@router.post(
    "/stores/{store_code}/simulated-sales-14d",
    response_model=RecentStoreSalesSimulationResponse,
    tags=["stores", "sales"],
)
def generate_recent_store_sales(
    store_code: str,
    payload: RecentStoreSalesSimulationRequest,
    authorization: Optional[str] = Header(default=None),
) -> RecentStoreSalesSimulationResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["store_code"] = store_code
    data["generated_by"] = current_user["username"]
    return RecentStoreSalesSimulationResponse(**state.generate_recent_store_sales(data))


@router.get("/stores/dispatch-bales", response_model=list[StoreDispatchBaleResponse], tags=["stores"])
def list_store_dispatch_bales(
    store_code: Optional[str] = Query(default=None),
    task_no: Optional[str] = Query(default=None),
    shipment_no: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    assigned_employee: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[StoreDispatchBaleResponse]:
    current_user = _require_current_user(authorization=authorization)
    if _is_store_clerk_user(current_user):
        store_code = str(current_user.get("store_code") or store_code or "").strip().upper() or None
        assigned_employee = str(current_user.get("username") or "").strip() or None
    rows = state.list_store_dispatch_bales(
        store_code=store_code,
        task_no=task_no,
        shipment_no=shipment_no,
        status=status,
        assigned_employee=assigned_employee,
    )
    return [StoreDispatchBaleResponse(**row) for row in rows]


@router.post("/stores/dispatch-bales/{bale_no}/accept", response_model=StoreDispatchBaleResponse, tags=["stores"])
def accept_store_dispatch_bale(
    bale_no: str,
    payload: StoreDispatchBaleAcceptRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreDispatchBaleResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["accepted_by"] = current_user["username"]
    return StoreDispatchBaleResponse(**state.accept_store_dispatch_bale(bale_no, data))


@router.post("/stores/dispatch-bales/{bale_no}/assign", response_model=StoreDispatchBaleResponse, tags=["stores"])
def assign_store_dispatch_bale(
    bale_no: str,
    payload: StoreDispatchBaleAssignRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreDispatchBaleResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["assigned_by"] = current_user["username"]
    return StoreDispatchBaleResponse(**state.assign_store_dispatch_bale(bale_no, data))


@router.get("/stores/dispatch-bales/{bale_no}/tokens", response_model=list[ItemBarcodeTokenResponse], tags=["stores"])
def list_store_dispatch_bale_tokens(
    bale_no: str,
    authorization: Optional[str] = Header(default=None),
) -> list[ItemBarcodeTokenResponse]:
    current_user = _require_current_user(authorization=authorization)
    bale = state.get_store_dispatch_bale(bale_no)
    _enforce_store_clerk_bale_access(current_user, bale)
    rows = state.get_store_dispatch_bale_tokens(bale_no)
    return [ItemBarcodeTokenResponse(**row) for row in rows]


@router.patch("/stores/item-barcode-tokens/{token_no}/edit", response_model=ItemBarcodeTokenResponse, tags=["stores"])
def edit_item_barcode_token_for_store(
    token_no: str,
    payload: ItemBarcodeTokenStoreEditRequest,
    authorization: Optional[str] = Header(default=None),
) -> ItemBarcodeTokenResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return ItemBarcodeTokenResponse(**state.update_item_barcode_token_store_edit(token_no, data))


@router.post("/warehouse/sorting-tasks/{task_no}/results", response_model=SortingTaskResponse, tags=["warehouse"])
def submit_sorting_task_results(
    task_no: str,
    payload: SortingTaskResultSubmit,
    authorization: Optional[str] = Header(default=None),
) -> SortingTaskResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return SortingTaskResponse(**state.submit_sorting_task_results(task_no, data))


@router.get("/warehouse/sorting-stock", response_model=list[SortingStockResponse], tags=["warehouse"])
def list_sorting_stock(
    authorization: Optional[str] = Header(default=None),
) -> list[SortingStockResponse]:
    _require_current_user(authorization=authorization)
    return [SortingStockResponse(**row) for row in state.list_sorting_stock()]


@router.patch(
    "/warehouse/sorting-stock/rack",
    response_model=SortingStockRackUpdateResponse,
    tags=["warehouse"],
)
def update_sorting_stock_rack(
    payload: SortingStockRackUpdateRequest,
    authorization: Optional[str] = Header(default=None),
) -> SortingStockRackUpdateResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["updated_by"] = current_user["username"]
    return SortingStockRackUpdateResponse(**state.update_sorting_stock_rack(data))


@router.get("/products", response_model=list[ProductResponse], tags=["products"])
def list_products(
    authorization: Optional[str] = Header(default=None),
) -> list[ProductResponse]:
    _require_current_user(authorization=authorization)
    return [ProductResponse(**product) for product in state.list_products()]


@router.post("/products/{product_id}/barcode", response_model=ProductResponse, tags=["products"])
def assign_product_barcode(
    product_id: int,
    payload: ProductBarcodeAssignRequest,
    authorization: Optional[str] = Header(default=None),
) -> ProductResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["assigned_by"] = current_user["username"]
    return ProductResponse(**state.assign_barcode_to_product(product_id, data))


@router.post(
    "/warehouse/receipts",
    response_model=GoodsReceiptResponse,
    tags=["warehouse"],
)
def create_goods_receipt(
    payload: GoodsReceiptCreate,
    authorization: Optional[str] = Header(default=None),
) -> GoodsReceiptResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return GoodsReceiptResponse(**state.create_goods_receipt(data))


@router.get(
    "/warehouse/receipts",
    response_model=list[GoodsReceiptResponse],
    tags=["warehouse"],
)
def list_goods_receipts(
    authorization: Optional[str] = Header(default=None),
) -> list[GoodsReceiptResponse]:
    _require_current_user(authorization=authorization)
    return [GoodsReceiptResponse(**receipt) for receipt in state.list_goods_receipts()]


@router.get(
    "/warehouse/stock",
    response_model=list[WarehouseStockResponse],
    tags=["warehouse"],
)
def list_warehouse_stock(
    authorization: Optional[str] = Header(default=None),
) -> list[WarehouseStockResponse]:
    _require_current_user(authorization=authorization)
    return [WarehouseStockResponse(**row) for row in state.list_warehouse_stock()]


@router.get(
    "/stores/stock",
    response_model=list[StoreStockResponse],
    tags=["stores"],
)
def list_store_stock(
    authorization: Optional[str] = Header(default=None),
) -> list[StoreStockResponse]:
    _require_current_user(authorization=authorization)
    return [StoreStockResponse(**row) for row in state.list_store_stock()]


@router.get(
    "/stores/{store_code}/stock/{barcode}",
    response_model=StoreStockLookupResponse,
    tags=["stores"],
)
def get_store_stock_lookup(
    store_code: str,
    barcode: str,
    authorization: Optional[str] = Header(default=None),
) -> StoreStockLookupResponse:
    _require_current_user(authorization=authorization)
    return StoreStockLookupResponse(**state.get_store_stock_lookup(store_code, barcode))


@router.post(
    "/stores/{store_code}/rack-locations/initialize",
    response_model=StoreRackInitializationResponse,
    tags=["stores"],
)
def initialize_store_rack_locations(
    store_code: str,
    authorization: Optional[str] = Header(default=None),
) -> StoreRackInitializationResponse:
    current_user = _require_current_user(authorization=authorization)
    racks = state.initialize_store_racks(
        store_code,
        STORE_RACK_TEMPLATE,
        initialized_by=current_user["username"],
    )
    rack_models = [StoreRackLocationResponse(**row) for row in racks]
    return StoreRackInitializationResponse(
        store_code=store_code,
        total_racks=len(rack_models),
        racks=rack_models,
    )


@router.get(
    "/stores/{store_code}/rack-locations",
    response_model=list[StoreRackLocationResponse],
    tags=["stores"],
)
def list_store_rack_locations(
    store_code: str,
    authorization: Optional[str] = Header(default=None),
) -> list[StoreRackLocationResponse]:
    _require_current_user(authorization=authorization)
    return [StoreRackLocationResponse(**row) for row in state.list_store_racks(store_code)]


@router.post(
    "/stores/{store_code}/rack-assignments",
    response_model=StoreRackAssignmentResponse,
    tags=["stores"],
)
def assign_store_rack(
    store_code: str,
    payload: StoreRackAssignmentRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreRackAssignmentResponse:
    current_user = _require_current_user(authorization=authorization)
    return StoreRackAssignmentResponse(
        **state.assign_store_rack(
            store_code=store_code,
            barcode=payload.barcode,
            rack_code=payload.rack_code,
            updated_by=current_user["username"],
        )
    )


@router.post(
    "/print-jobs/bale-label",
    response_model=BaleLabelPrintJobResponse,
    tags=["printing"],
)
def create_bale_label_print_station_job(
    payload: BaleLabelPrintJobCreate,
    authorization: Optional[str] = Header(default=None),
) -> BaleLabelPrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return BaleLabelPrintJobResponse(**state.create_bale_label_print_station_job(data))


@router.get(
    "/print-jobs/pending",
    response_model=list[BaleLabelPrintJobResponse],
    tags=["printing"],
)
def list_pending_bale_label_print_station_jobs(
    station_id: str = Query(default=""),
    authorization: Optional[str] = Header(default=None),
) -> list[BaleLabelPrintJobResponse]:
    _require_current_user(authorization=authorization)
    jobs = state.list_pending_print_station_jobs(station_id=station_id)
    return [BaleLabelPrintJobResponse(**job) for job in jobs]


@router.post(
    "/print-jobs/{job_id}/claim",
    response_model=BaleLabelPrintJobResponse,
    tags=["printing"],
)
def claim_bale_label_print_station_job(
    job_id: int,
    payload: PrintStationClaimRequest,
    authorization: Optional[str] = Header(default=None),
) -> BaleLabelPrintJobResponse:
    _require_current_user(authorization=authorization)
    return BaleLabelPrintJobResponse(**state.claim_print_station_job(job_id, station_id=payload.station_id))


@router.post(
    "/print-jobs/labels",
    response_model=PrintJobResponse,
    tags=["printing"],
)
def create_label_print_job(
    payload: LabelPrintJobCreate,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return PrintJobResponse(**state.create_label_print_job(data))


@router.post(
    "/print-jobs/item-tokens",
    response_model=list[PrintJobResponse],
    tags=["printing"],
)
def queue_item_token_print_jobs(
    payload: ItemBarcodeTokenPrintJobCreate,
    authorization: Optional[str] = Header(default=None),
) -> list[PrintJobResponse]:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    jobs = state.queue_item_barcode_token_print_jobs(data)
    return [PrintJobResponse(**job) for job in jobs]


@router.post(
    "/warehouse/receipts/{receipt_no}/print-jobs/labels",
    response_model=list[PrintJobResponse],
    tags=["printing", "warehouse"],
)
def queue_receipt_label_print_jobs(
    receipt_no: str,
    authorization: Optional[str] = Header(default=None),
) -> list[PrintJobResponse]:
    current_user = _require_current_user(authorization=authorization)
    jobs = state.queue_receipt_label_print_jobs(receipt_no, requested_by=current_user["username"])
    return [PrintJobResponse(**job) for job in jobs]


@router.get(
    "/transfers/recommendations",
    response_model=list[TransferRecommendationResponse],
    tags=["transfers"],
)
def list_transfer_recommendations(
    authorization: Optional[str] = Header(default=None),
) -> list[TransferRecommendationResponse]:
    _require_current_user(authorization=authorization)
    return [TransferRecommendationResponse(**row) for row in state.list_transfer_recommendations()]


@router.post(
    "/transfers/recommendations",
    response_model=TransferRecommendationResponse,
    tags=["transfers"],
)
def create_transfer_recommendation(
    payload: TransferRecommendationRequest,
    authorization: Optional[str] = Header(default=None),
) -> TransferRecommendationResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return TransferRecommendationResponse(**state.create_transfer_recommendation(data))


@router.post(
    "/transfers/recommendations/{recommendation_no}/create-transfer",
    response_model=TransferOrderResponse,
    tags=["transfers"],
)
def create_transfer_from_recommendation(
    recommendation_no: str,
    payload: RecommendationTransferCreateRequest,
    authorization: Optional[str] = Header(default=None),
) -> TransferOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return TransferOrderResponse(**state.create_transfer_from_recommendation(recommendation_no, data))


@router.get(
    "/print-jobs",
    response_model=list[PrintJobResponse],
    tags=["printing"],
)
def list_print_jobs(
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[PrintJobResponse]:
    _require_current_user(authorization=authorization)
    return [PrintJobResponse(**job) for job in state.list_print_jobs(status=status)]


@router.get(
    "/system/printers",
    response_model=list[SystemPrinterResponse],
    tags=["printing"],
)
def list_system_printers(
    authorization: Optional[str] = Header(default=None),
) -> list[SystemPrinterResponse]:
    _require_current_user(authorization=authorization)
    return [SystemPrinterResponse(**row) for row in _list_system_printers()]


@router.get(
    "/print-jobs/{job_id}",
    response_model=PrintJobResponse,
    tags=["printing"],
)
def get_print_job(
    job_id: int,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse:
    _require_current_user(authorization=authorization)
    return PrintJobResponse(**state.get_print_job(job_id))


@router.get(
    "/print-jobs/{job_id}/preview",
    response_class=HTMLResponse,
    tags=["printing"],
)
def preview_print_job(
    job_id: int,
    authorization: Optional[str] = Header(default=None),
    access_token: Optional[str] = Query(default=None),
    autoprint: Optional[int] = Query(default=0),
    template_code: Optional[str] = Query(default=None),
) -> HTMLResponse:
    _require_current_user(authorization=authorization, access_token=access_token)
    job = state.get_print_job(job_id)
    if job["job_type"] in {"barcode_label", "bale_barcode_label", "item_token_label"}:
        payload = _merge_print_payload_with_template(job, template_code)
        return HTMLResponse(_build_barcode_preview_html(job, payload, autoprint=bool(autoprint)))

    payload = job["print_payload"]
    rows = "".join(
        [
            (
                f"<tr><td style='border:1px solid #222;padding:8px;'>{item['barcode']}</td>"
                f"<td style='border:1px solid #222;padding:8px;'>{item['product_name']}</td>"
                f"<td style='border:1px solid #222;padding:8px;text-align:right;'>{item['requested_qty']}</td>"
                f"<td style='border:1px solid #222;padding:8px;text-align:right;'>{item['approved_qty']}</td>"
                f"<td style='border:1px solid #222;padding:8px;text-align:right;'>{item['received_qty']}</td></tr>"
            )
            for item in payload["items"]
        ]
    )
    auto_print_script = (
        "<script>window.addEventListener('load',()=>setTimeout(()=>window.print(),200));</script>"
        if autoprint
        else ""
    )
    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; padding: 24px;">
        <h2>Transfer Order {payload['transfer_no']}</h2>
        <p>Warehouse: {payload['from_warehouse_code']} -> Store: {payload['to_store_code']}</p>
        <p>Status: {payload['status']}</p>
        <table style="border-collapse: collapse; width: 100%;">
          <thead>
            <tr>
              <th style="border:1px solid #222;padding:8px;">Barcode</th>
              <th style="border:1px solid #222;padding:8px;">Product</th>
              <th style="border:1px solid #222;padding:8px;">Requested</th>
              <th style="border:1px solid #222;padding:8px;">Approved</th>
              <th style="border:1px solid #222;padding:8px;">Received</th>
            </tr>
          </thead>
          <tbody>{rows}</tbody>
        </table>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:24px;margin-top:32px;">
          <div><div>Warehouse Picker</div><div style="border-bottom:1px solid #222;height:40px;"></div></div>
          <div><div>Dispatcher</div><div style="border-bottom:1px solid #222;height:40px;"></div></div>
          <div><div>Store Manager</div><div style="border-bottom:1px solid #222;height:40px;"></div></div>
        </div>
        {auto_print_script}
      </body>
    </html>
    """
    return HTMLResponse(html)


@router.post(
    "/print-jobs/bale-direct/print",
    response_model=MessageResponse,
    tags=["printing"],
)
def direct_print_bale_label(
    payload: BaleDirectPrintRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    _require_current_user(authorization=authorization)
    selected_printer = str(payload.printer_name or "").strip()
    if not selected_printer:
        raise HTTPException(status_code=400, detail="请先选择打印机。")
    printers = _list_system_printers()
    printer_row = _find_system_printer(printers, selected_printer)
    if not printer_row:
        raise HTTPException(status_code=404, detail=f"系统里没有找到打印机 {selected_printer}")
    printer_destination = _resolve_printer_destination(printer_row, selected_printer)
    synthetic_job = {
        "job_type": "bale_barcode_label",
        "label_size": payload.template_code.replace("bale_", "").replace("_", "x"),
        "copies": payload.copies,
    }
    merged_payload = {
        "barcode_value": payload.barcode_value,
        "scan_token": payload.scan_token,
        "bale_barcode": payload.bale_barcode,
        "legacy_bale_barcode": payload.legacy_bale_barcode,
        "supplier_name": payload.supplier_name,
        "category_main": payload.category_main,
        "category_sub": payload.category_sub,
        "category_display": payload.category_display,
        "package_position_label": payload.package_position_label,
        "serial_no": payload.serial_no,
        "total_packages": payload.total_packages,
        "shipment_no": payload.shipment_no,
        "parcel_batch_no": payload.parcel_batch_no,
        "unload_date": payload.unload_date,
        "template_scope": payload.template_scope,
        "store_name": payload.store_name,
        "transfer_order_no": payload.transfer_order_no,
        "bale_piece_summary": payload.bale_piece_summary,
        "total_quantity": payload.total_quantity,
        "packing_list": payload.packing_list,
        "dispatch_bale_no": payload.dispatch_bale_no,
        "outbound_time": payload.outbound_time,
        "status": payload.status,
        "cat": payload.cat,
        "sub": payload.sub,
        "grade": payload.grade,
        "qty": payload.qty,
        "weight": payload.weight,
        "code": payload.code,
    }
    merged_payload = _merge_print_payload_with_template(
        {"print_payload": merged_payload, "label_size": synthetic_job["label_size"]},
        payload.template_code,
    )
    if _is_tspl_raw_printer(printer_row):
        raw_tspl = _build_tspl_barcode_payload(synthetic_job, merged_payload)
        _send_raw_print_job(printer_destination, raw_tspl, copies=max(int(payload.copies or 1), 1))
        return MessageResponse(message=f"已通过 TSPL 原始指令发送到打印机 {selected_printer}，请确认标签是否开始出纸。")
    with NamedTemporaryFile(prefix="bale-label-", suffix=".pdf", delete=False) as handle:
        pdf_path = handle.name
    try:
        _render_barcode_print_job_pdf(synthetic_job, merged_payload, pdf_path)
        print_result = subprocess.run(
            ["/usr/bin/lp", "-d", printer_destination, "-n", str(max(int(payload.copies or 1), 1)), pdf_path],
            capture_output=True,
            text=True,
            check=False,
        )
        if print_result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=(print_result.stderr or print_result.stdout or "打印任务发送失败").strip(),
            )
    finally:
        try:
            os.remove(pdf_path)
        except OSError:
            pass
    return MessageResponse(message=f"已发送到打印机 {selected_printer}，请确认标签是否开始出纸。")


@router.post(
    "/print-jobs/bale-direct/print-batch",
    response_model=MessageResponse,
    tags=["printing"],
)
def direct_print_bale_label_batch(
    payload: BaleDirectPrintBatchRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    _require_current_user(authorization=authorization)
    selected_printer = str(payload.printer_name or "").strip()
    if not selected_printer:
        raise HTTPException(status_code=400, detail="请先选择打印机。")
    printers = _list_system_printers()
    printer_row = _find_system_printer(printers, selected_printer)
    if not printer_row:
        raise HTTPException(status_code=404, detail=f"系统里没有找到打印机 {selected_printer}")
    printer_destination = _resolve_printer_destination(printer_row, selected_printer)
    synthetic_jobs = [
        {
            "job_type": "bale_barcode_label",
            "label_size": payload.template_code.replace("bale_", "").replace("_", "x"),
            "copies": item.copies,
        }
        for item in payload.items
    ]
    merged_payloads = [
        _merge_print_payload_with_template({
            "print_payload": {
            "barcode_value": item.barcode_value,
            "scan_token": item.scan_token,
            "bale_barcode": item.bale_barcode,
            "legacy_bale_barcode": item.legacy_bale_barcode,
            "supplier_name": item.supplier_name,
            "category_main": item.category_main,
            "category_sub": item.category_sub,
            "category_display": item.category_display,
            "package_position_label": item.package_position_label,
            "serial_no": item.serial_no,
            "total_packages": item.total_packages,
            "shipment_no": item.shipment_no,
            "parcel_batch_no": item.parcel_batch_no,
            "unload_date": item.unload_date,
            "template_scope": item.template_scope,
            "store_name": item.store_name,
            "transfer_order_no": item.transfer_order_no,
            "bale_piece_summary": item.bale_piece_summary,
            "total_quantity": item.total_quantity,
            "packing_list": item.packing_list,
            "dispatch_bale_no": item.dispatch_bale_no,
            "outbound_time": item.outbound_time,
            "status": item.status,
            "cat": item.cat,
            "sub": item.sub,
            "grade": item.grade,
            "qty": item.qty,
            "weight": item.weight,
            "code": item.code,
            "copies": item.copies,
            },
            "label_size": payload.template_code.replace("bale_", "").replace("_", "x"),
        }, payload.template_code)
        for item in payload.items
    ]
    if _is_tspl_raw_printer(printer_row):
        raw_tspl = _build_tspl_barcode_batch_payload(synthetic_jobs, merged_payloads)
        _send_raw_print_job(printer_destination, raw_tspl, copies=1)
        return MessageResponse(
            message=f"已通过 TSPL 原始指令一次性发送 {len(payload.items)} 张到打印机 {selected_printer}，请确认标签是否连续出纸。"
        )
    for job, merged_payload in zip(synthetic_jobs, merged_payloads):
        direct_print_bale_label(
            BaleDirectPrintRequest(
                printer_name=selected_printer,
                template_code=payload.template_code,
                copies=max(int(merged_payload.get("copies") or 1), 1),
                barcode_value=str(merged_payload.get("barcode_value") or "").strip(),
                scan_token=str(merged_payload.get("scan_token") or "").strip(),
                bale_barcode=str(merged_payload.get("bale_barcode") or "").strip(),
                legacy_bale_barcode=str(merged_payload.get("legacy_bale_barcode") or "").strip(),
                supplier_name=str(merged_payload.get("supplier_name") or "").strip(),
                category_main=str(merged_payload.get("category_main") or "").strip(),
                category_sub=str(merged_payload.get("category_sub") or "").strip(),
                category_display=str(merged_payload.get("category_display") or "").strip(),
                package_position_label=str(merged_payload.get("package_position_label") or "").strip(),
                serial_no=int(merged_payload.get("serial_no") or 0),
                total_packages=int(merged_payload.get("total_packages") or 0),
                shipment_no=str(merged_payload.get("shipment_no") or "").strip(),
                parcel_batch_no=str(merged_payload.get("parcel_batch_no") or "").strip(),
                unload_date=str(merged_payload.get("unload_date") or "").strip(),
                template_scope=str(merged_payload.get("template_scope") or "").strip(),
                store_name=str(merged_payload.get("store_name") or "").strip(),
                transfer_order_no=str(merged_payload.get("transfer_order_no") or "").strip(),
                bale_piece_summary=str(merged_payload.get("bale_piece_summary") or "").strip(),
                total_quantity=str(merged_payload.get("total_quantity") or "").strip(),
                packing_list=str(merged_payload.get("packing_list") or "").strip(),
                dispatch_bale_no=str(merged_payload.get("dispatch_bale_no") or "").strip(),
                outbound_time=str(merged_payload.get("outbound_time") or "").strip(),
                status=str(merged_payload.get("status") or "").strip(),
                cat=str(merged_payload.get("cat") or "").strip(),
                sub=str(merged_payload.get("sub") or "").strip(),
                grade=str(merged_payload.get("grade") or "").strip(),
                qty=str(merged_payload.get("qty") or "").strip(),
                weight=str(merged_payload.get("weight") or "").strip(),
                code=str(merged_payload.get("code") or "").strip(),
            ),
            authorization=authorization,
        )
    return MessageResponse(message=f"已顺序发送 {len(payload.items)} 张到打印机 {selected_printer}，请确认标签是否连续出纸。")


@router.post(
    "/print-jobs/candidate-lab/print",
    response_model=MessageResponse,
    tags=["printing"],
)
def direct_print_candidate_lab_label(
    payload: LabelCandidatePrintRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    _require_current_user(authorization=authorization)
    selected_printer = str(payload.printer_name or "").strip()
    if not selected_printer:
        raise HTTPException(status_code=400, detail="请先选择打印机。")
    printers = _list_system_printers()
    printer_row = _find_system_printer(printers, selected_printer)
    if not printer_row:
        raise HTTPException(status_code=404, detail=f"系统里没有找到打印机 {selected_printer}")
    if not _is_tspl_raw_printer(printer_row):
        raise HTTPException(status_code=400, detail="候选页测试只支持 Deli TSPL 直打。")
    printer_destination = _resolve_printer_destination(printer_row, selected_printer)
    raw_tspl = _build_candidate_lab_tspl_payload(payload.model_dump())
    _send_raw_print_job(printer_destination, raw_tspl, copies=1)
    return MessageResponse(message=f"已把候选版 {payload.candidate_id} 发送到 {selected_printer}，请确认是否正常出纸。")


@router.post(
    "/print-jobs/candidate-lab/print-batch",
    response_model=MessageResponse,
    tags=["printing"],
)
def direct_print_candidate_lab_batch(
    payload: LabelCandidateBatchPrintRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    _require_current_user(authorization=authorization)
    selected_printer = str(payload.printer_name or "").strip()
    if not selected_printer:
        raise HTTPException(status_code=400, detail="请先选择打印机。")
    printers = _list_system_printers()
    printer_row = _find_system_printer(printers, selected_printer)
    if not printer_row:
        raise HTTPException(status_code=404, detail=f"系统里没有找到打印机 {selected_printer}")
    if not _is_tspl_raw_printer(printer_row):
        raise HTTPException(status_code=400, detail="候选页测试只支持 Deli TSPL 直打。")
    printer_destination = _resolve_printer_destination(printer_row, selected_printer)
    raw_tspl = _build_candidate_lab_tspl_batch_payload(payload.model_dump())
    _send_raw_print_job(printer_destination, raw_tspl, copies=1)
    return MessageResponse(message=f"已把 {len(payload.candidates)} 张候选版连续发送到 {selected_printer}，请确认是否连续出纸。")


@router.post(
    "/print-jobs/{job_id}/direct-print",
    response_model=MessageResponse,
    tags=["printing"],
)
def direct_print_job(
    job_id: int,
    authorization: Optional[str] = Header(default=None),
    printer_name: Optional[str] = Query(default=None),
    template_code: Optional[str] = Query(default=None),
) -> MessageResponse:
    _require_current_user(authorization=authorization)
    job = state.get_print_job(job_id)
    if job["job_type"] not in {"barcode_label", "bale_barcode_label", "item_token_label"}:
        raise HTTPException(status_code=400, detail="当前只支持直接打印条码标签。")
    payload = _merge_print_payload_with_template(job, template_code)
    selected_printer = str(printer_name or job.get("printer_name") or "").strip()
    if not selected_printer:
        raise HTTPException(status_code=400, detail="请先选择打印机。")
    printers = _list_system_printers()
    printer_row = _find_system_printer(printers, selected_printer)
    if not printer_row:
        raise HTTPException(status_code=404, detail=f"系统里没有找到打印机 {selected_printer}")
    printer_destination = _resolve_printer_destination(printer_row, selected_printer)
    if _is_tspl_raw_printer(printer_row):
        raw_tspl = _build_tspl_barcode_payload(job, payload)
        _send_raw_print_job(printer_destination, raw_tspl, copies=max(int(job.get("copies") or 1), 1))
        return MessageResponse(message=f"已通过 TSPL 原始指令发送到打印机 {selected_printer}，请确认标签是否开始出纸。")
    with NamedTemporaryFile(prefix="bale-label-", suffix=".pdf", delete=False) as handle:
        pdf_path = handle.name
    try:
        _render_barcode_print_job_pdf(job, payload, pdf_path)
        print_result = subprocess.run(
            ["/usr/bin/lp", "-d", printer_destination, "-n", str(max(int(job.get("copies") or 1), 1)), pdf_path],
            capture_output=True,
            text=True,
            check=False,
        )
        if print_result.returncode != 0:
            raise HTTPException(
                status_code=500,
                detail=(print_result.stderr or print_result.stdout or "打印任务发送失败").strip(),
            )
    finally:
        try:
            os.remove(pdf_path)
        except OSError:
            pass
    return MessageResponse(message=f"已发送到打印机 {selected_printer}，请在现场确认标签已真实打印。")


@router.post("/transfers", response_model=TransferOrderResponse, tags=["transfers"])
def create_transfer_order(
    payload: TransferOrderCreate,
    authorization: Optional[str] = Header(default=None),
) -> TransferOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return TransferOrderResponse(**state.create_transfer_order(data))


@router.get("/transfers", response_model=list[TransferOrderResponse], tags=["transfers"])
def list_transfer_orders(
    authorization: Optional[str] = Header(default=None),
) -> list[TransferOrderResponse]:
    _require_current_user(authorization=authorization)
    return [TransferOrderResponse(**order) for order in state.list_transfer_orders()]


@router.get("/transfers/{transfer_no}", response_model=TransferOrderResponse, tags=["transfers"])
def get_transfer_order(
    transfer_no: str,
    authorization: Optional[str] = Header(default=None),
) -> TransferOrderResponse:
    _require_current_user(authorization=authorization)
    return TransferOrderResponse(**state.get_transfer_order(transfer_no))


@router.post("/picking-waves", response_model=PickingWaveResponse, tags=["transfers"])
def create_picking_wave(
    payload: PickingWaveCreate,
    authorization: Optional[str] = Header(default=None),
) -> PickingWaveResponse:
    _require_current_user(authorization=authorization)
    return PickingWaveResponse(**state.create_picking_wave(payload.model_dump()))


@router.get("/picking-waves", response_model=list[PickingWaveResponse], tags=["transfers"])
def list_picking_waves(
    authorization: Optional[str] = Header(default=None),
) -> list[PickingWaveResponse]:
    _require_current_user(authorization=authorization)
    return [PickingWaveResponse(**row) for row in state.list_picking_waves()]


@router.get("/picking-waves/{wave_no}", response_model=PickingWaveResponse, tags=["transfers"])
def get_picking_wave(
    wave_no: str,
    authorization: Optional[str] = Header(default=None),
) -> PickingWaveResponse:
    _require_current_user(authorization=authorization)
    return PickingWaveResponse(**state.get_picking_wave(wave_no))


@router.get(
    "/transfers/{transfer_no}/store-delivery-execution-orders",
    response_model=list[StoreDeliveryExecutionOrderResponse],
    tags=["transfers"],
)
def list_store_delivery_execution_orders(
    transfer_no: str,
    authorization: Optional[str] = Header(default=None),
) -> list[StoreDeliveryExecutionOrderResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_store_delivery_execution_orders(transfer_no=transfer_no)
    return [StoreDeliveryExecutionOrderResponse(**row) for row in rows]


@router.post(
    "/transfers/{transfer_no}/store-delivery-execution-orders",
    response_model=StoreDeliveryExecutionOrderResponse,
    tags=["transfers"],
)
def create_store_delivery_execution_order(
    transfer_no: str,
    payload: StoreDeliveryExecutionOrderCreateRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreDeliveryExecutionOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return StoreDeliveryExecutionOrderResponse(**state.create_store_delivery_execution_order(transfer_no, data))


@router.get(
    "/receiving-sessions",
    response_model=list[ReceivingSessionResponse],
    tags=["receiving"],
)
def list_all_receiving_sessions(
    authorization: Optional[str] = Header(default=None),
) -> list[ReceivingSessionResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_transfer_receiving_sessions()
    return [ReceivingSessionResponse(**row) for row in rows]


@router.get(
    "/store-token-receiving-sessions",
    response_model=list[StoreTokenReceivingSessionResponse],
    tags=["receiving", "stores"],
)
def list_store_token_receiving_sessions(
    store_code: Optional[str] = Query(default=None),
    task_no: Optional[str] = Query(default=None),
    shipment_no: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[StoreTokenReceivingSessionResponse]:
    current_user = _require_current_user(authorization=authorization)
    if _is_store_clerk_user(current_user):
        store_code = str(current_user.get("store_code") or store_code or "").strip().upper() or None
    rows = state.list_store_token_receiving_sessions(store_code=store_code, task_no=task_no, shipment_no=shipment_no)
    if _is_store_clerk_user(current_user):
        username = str(current_user.get("username") or "").strip().lower()
        rows = [row for row in rows if str(row.get("assigned_employee") or "").strip().lower() == username]
    return [StoreTokenReceivingSessionResponse(**row) for row in rows]


@router.post(
    "/store-token-receiving-sessions/start",
    response_model=StoreTokenReceivingSessionResponse,
    tags=["receiving", "stores"],
)
def start_store_token_receiving_session(
    payload: StoreTokenReceivingSessionStartRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreTokenReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["started_by"] = current_user["username"]
    return StoreTokenReceivingSessionResponse(**state.start_store_token_receiving_session(data))


@router.get(
    "/store-token-receiving-sessions/{session_no}",
    response_model=StoreTokenReceivingSessionResponse,
    tags=["receiving", "stores"],
)
def get_store_token_receiving_session(
    session_no: str,
    authorization: Optional[str] = Header(default=None),
) -> StoreTokenReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    session = state.get_store_token_receiving_session(session_no)
    _enforce_store_clerk_session_access(current_user, session)
    return StoreTokenReceivingSessionResponse(**session)


@router.get(
    "/store-token-receiving-sessions/{session_no}/placement-suggestions/{token_no}",
    response_model=StoreTokenPlacementSuggestionResponse,
    tags=["receiving", "stores"],
)
def get_store_token_receiving_session_placement_suggestion(
    session_no: str,
    token_no: str,
    authorization: Optional[str] = Header(default=None),
) -> StoreTokenPlacementSuggestionResponse:
    current_user = _require_current_user(authorization=authorization)
    session = state.get_store_token_receiving_session(session_no)
    _enforce_store_clerk_session_access(current_user, session)
    return StoreTokenPlacementSuggestionResponse(
        **state.get_store_token_placement_suggestion(session_no, token_no)
    )


@router.post(
    "/store-token-receiving-sessions/{session_no}/batches",
    response_model=StoreTokenReceivingSessionResponse,
    tags=["receiving", "stores"],
)
def add_store_token_receiving_batch(
    session_no: str,
    payload: StoreTokenReceivingBatchAddRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreTokenReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["recorded_by"] = current_user["username"]
    return StoreTokenReceivingSessionResponse(**state.add_store_token_receiving_batch(session_no, data))


@router.post(
    "/store-token-receiving-sessions/{session_no}/finalize",
    response_model=StoreTokenReceivingSessionResponse,
    tags=["receiving", "stores"],
)
def finalize_store_token_receiving_session(
    session_no: str,
    payload: ReceivingSessionFinalizeRequest,
    authorization: Optional[str] = Header(default=None),
) -> StoreTokenReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["finalized_by"] = current_user["username"]
    return StoreTokenReceivingSessionResponse(**state.finalize_store_token_receiving_session(session_no, data))


@router.get(
    "/transfers/{transfer_no}/receiving-sessions",
    response_model=list[ReceivingSessionResponse],
    tags=["receiving", "transfers"],
)
def list_transfer_receiving_sessions(
    transfer_no: str,
    authorization: Optional[str] = Header(default=None),
) -> list[ReceivingSessionResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_transfer_receiving_sessions(transfer_no=transfer_no)
    return [ReceivingSessionResponse(**row) for row in rows]


@router.post(
    "/transfers/{transfer_no}/receiving-sessions/start",
    response_model=ReceivingSessionResponse,
    tags=["receiving", "transfers"],
)
def start_transfer_receiving_session(
    transfer_no: str,
    payload: ReceivingSessionStartRequest,
    authorization: Optional[str] = Header(default=None),
) -> ReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["started_by"] = current_user["username"]
    return ReceivingSessionResponse(**state.start_transfer_receiving_session(transfer_no, data))


@router.get(
    "/receiving-sessions/{session_no}",
    response_model=ReceivingSessionResponse,
    tags=["receiving"],
)
def get_receiving_session(
    session_no: str,
    authorization: Optional[str] = Header(default=None),
) -> ReceivingSessionResponse:
    _require_current_user(authorization=authorization)
    return ReceivingSessionResponse(**state.get_transfer_receiving_session(session_no))


@router.get(
    "/receiving-sessions/{session_no}/placement-suggestions/{barcode}",
    response_model=PlacementSuggestionResponse,
    tags=["receiving", "stores"],
)
def get_receiving_session_placement_suggestion(
    session_no: str,
    barcode: str,
    authorization: Optional[str] = Header(default=None),
) -> PlacementSuggestionResponse:
    _require_current_user(authorization=authorization)
    return PlacementSuggestionResponse(
        **state.get_receiving_session_placement_suggestion(session_no, barcode)
    )


@router.post(
    "/receiving-sessions/{session_no}/batches",
    response_model=ReceivingSessionResponse,
    tags=["receiving"],
)
def add_receiving_session_batch(
    session_no: str,
    payload: ReceivingBatchAddRequest,
    authorization: Optional[str] = Header(default=None),
) -> ReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["recorded_by"] = current_user["username"]
    return ReceivingSessionResponse(**state.add_receiving_session_batch(session_no, data))


@router.post(
    "/receiving-sessions/{session_no}/finalize",
    response_model=ReceivingSessionResponse,
    tags=["receiving"],
)
def finalize_receiving_session(
    session_no: str,
    payload: ReceivingSessionFinalizeRequest,
    authorization: Optional[str] = Header(default=None),
) -> ReceivingSessionResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["finalized_by"] = current_user["username"]
    return ReceivingSessionResponse(**state.finalize_transfer_receiving_session(session_no, data))


@router.post("/returns", response_model=ReturnOrderResponse, tags=["returns"])
def create_return_order(
    payload: ReturnOrderCreate,
    authorization: Optional[str] = Header(default=None),
) -> ReturnOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return ReturnOrderResponse(**state.create_return_order(data))


@router.get(
    "/stores/{store_code}/return-candidates",
    response_model=list[ReturnCandidateResponse],
    tags=["returns", "stores"],
)
def list_return_candidates(
    store_code: str,
    authorization: Optional[str] = Header(default=None),
) -> list[ReturnCandidateResponse]:
    _require_current_user(authorization=authorization)
    return [ReturnCandidateResponse(**row) for row in state.list_return_candidates(store_code)]


@router.post("/returns/from-selection", response_model=ReturnOrderResponse, tags=["returns"])
def create_return_order_from_selection(
    payload: ReturnSelectionCreate,
    authorization: Optional[str] = Header(default=None),
) -> ReturnOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["created_by"] = current_user["username"]
    return ReturnOrderResponse(**state.create_return_order_from_selection(data))


@router.get("/returns", response_model=list[ReturnOrderResponse], tags=["returns"])
def list_return_orders(
    authorization: Optional[str] = Header(default=None),
) -> list[ReturnOrderResponse]:
    _require_current_user(authorization=authorization)
    return [ReturnOrderResponse(**order) for order in state.list_return_orders()]


@router.get("/returns/{return_no}", response_model=ReturnOrderResponse, tags=["returns"])
def get_return_order(
    return_no: str,
    authorization: Optional[str] = Header(default=None),
) -> ReturnOrderResponse:
    _require_current_user(authorization=authorization)
    return ReturnOrderResponse(**state.get_return_order(return_no))


@router.post("/returns/{return_no}/dispatch", response_model=MessageResponse, tags=["returns"])
def dispatch_return_order(
    return_no: str,
    payload: ReturnOrderDispatchRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["dispatched_by"] = current_user["username"]
    order = state.dispatch_return_order(return_no, data)
    return MessageResponse(
        message=f"Return {order['return_no']} dispatched by {current_user['username']}"
    )


@router.post("/returns/{return_no}/receive", response_model=MessageResponse, tags=["returns"])
def receive_return_order(
    return_no: str,
    payload: ReturnOrderReceiveRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["received_by"] = current_user["username"]
    order = state.receive_return_order(return_no, data)
    return MessageResponse(
        message=f"Return {order['return_no']} received by {current_user['username']} into {order['ret_rack_code']}"
    )


@router.post(
    "/transfers/{transfer_no}/approve",
    response_model=TransferOrderResponse,
    tags=["transfers"],
)
def approve_transfer_order(
    transfer_no: str,
    payload: TransferApprovalRequest,
    authorization: Optional[str] = Header(default=None),
) -> TransferOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["approved_by"] = current_user["username"]
    return TransferOrderResponse(**state.approve_transfer_order(transfer_no, data))


@router.post(
    "/transfers/{transfer_no}/ship",
    response_model=TransferOrderResponse,
    tags=["transfers"],
)
def ship_transfer_order(
    transfer_no: str,
    payload: TransferShipRequest,
    authorization: Optional[str] = Header(default=None),
) -> TransferOrderResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["shipped_by"] = current_user["username"]
    return TransferOrderResponse(**state.ship_transfer_order(transfer_no, data))


@router.post(
    "/transfers/{transfer_no}/receive",
    response_model=MessageResponse,
    tags=["transfers"],
)
def receive_transfer_order(
    transfer_no: str,
    payload: TransferReceiveRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["receiver_name"] = current_user["username"]
    order = state.receive_transfer_order(transfer_no, data)
    return MessageResponse(
        message=(
            f"Transfer {order['transfer_no']} received by {current_user['username']} "
            f"with status {order['status']}"
        )
    )


@router.post(
    "/transfers/{transfer_no}/discrepancy-approval",
    response_model=MessageResponse,
    tags=["transfers"],
)
def approve_transfer_discrepancy(
    transfer_no: str,
    payload: DiscrepancyApprovalRequest,
    authorization: Optional[str] = Header(default=None),
) -> MessageResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["approved_by"] = current_user["username"]
    order = state.approve_transfer_discrepancy(transfer_no, data)
    result = order["discrepancy_approval_status"]
    return MessageResponse(
        message=f"Transfer {transfer_no} discrepancy {result} by {current_user['username']}"
    )


@router.get(
    "/inventory-adjustments",
    response_model=list[InventoryAdjustmentResponse],
    tags=["inventory"],
)
def list_inventory_adjustments(
    authorization: Optional[str] = Header(default=None),
) -> list[InventoryAdjustmentResponse]:
    _require_current_user(authorization=authorization)
    return [InventoryAdjustmentResponse(**row) for row in state.list_inventory_adjustments()]


@router.get(
    "/inventory-movements",
    response_model=list[InventoryMovementResponse],
    tags=["inventory"],
)
def list_inventory_movements(
    barcode: Optional[str] = Query(default=None),
    location_code: Optional[str] = Query(default=None),
    movement_type: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[InventoryMovementResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_inventory_movements(
        barcode=barcode,
        location_code=location_code,
        movement_type=movement_type,
    )
    return [InventoryMovementResponse(**row) for row in rows]


@router.get("/audit-events", response_model=list[AuditEventResponse], tags=["audit"])
def list_audit_events(
    authorization: Optional[str] = Header(default=None),
) -> list[AuditEventResponse]:
    _require_current_user(authorization=authorization)
    return [AuditEventResponse(**row) for row in state.list_audit_events()]


@router.get("/payments/anomalies", response_model=list[PaymentAnomalyResponse], tags=["payments"])
def list_payment_anomalies(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    anomaly_type: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[PaymentAnomalyResponse]:
    _require_current_user(authorization=authorization)
    rows = state.list_payment_anomalies(
        store_code=store_code,
        status=status,
        anomaly_type=anomaly_type,
    )
    return [PaymentAnomalyResponse(**row) for row in rows]


@router.post(
    "/payments/anomalies/{anomaly_no}/resolve",
    response_model=PaymentAnomalyResponse,
    tags=["payments"],
)
def resolve_payment_anomaly(
    anomaly_no: str,
    payload: PaymentAnomalyResolveRequest,
    authorization: Optional[str] = Header(default=None),
) -> PaymentAnomalyResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["resolved_by"] = current_user["username"]
    return PaymentAnomalyResponse(**state.resolve_payment_anomaly(anomaly_no, data))


@router.post(
    "/integrations/mpesa/import",
    response_model=list[MpesaCollectionResponse],
    tags=["integrations", "payments"],
)
def import_mpesa_collections(
    payload: MpesaCollectionImportRequest,
    authorization: Optional[str] = Header(default=None),
) -> list[MpesaCollectionResponse]:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["imported_by"] = current_user["username"]
    return [MpesaCollectionResponse(**row) for row in state.import_mpesa_collections(data)]


@router.get(
    "/integrations/mpesa/collections",
    response_model=list[MpesaCollectionResponse],
    tags=["integrations", "payments"],
)
def list_mpesa_collections(
    store_code: Optional[str] = Query(default=None),
    match_status: Optional[str] = Query(default=None),
    customer_id: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[MpesaCollectionResponse]:
    _require_current_user(authorization=authorization)
    return [
        MpesaCollectionResponse(**row)
        for row in state.list_mpesa_collections(
            store_code=store_code,
            match_status=match_status,
            customer_id=customer_id,
        )
    ]


@router.get(
    "/integrations/mpesa/customer-insights",
    response_model=list[MpesaCustomerInsightResponse],
    tags=["integrations", "payments"],
)
def list_mpesa_customer_insights(
    store_code: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[MpesaCustomerInsightResponse]:
    _require_current_user(authorization=authorization)
    return [
        MpesaCustomerInsightResponse(**row)
        for row in state.list_mpesa_customer_insights(store_code=store_code)
    ]


@router.post(
    "/integrations/mpesa/safaricom-callback",
    response_model=MpesaCallbackResponse,
    tags=["integrations", "payments"],
)
def ingest_safaricom_callback(
    payload: dict[str, Any],
) -> MpesaCallbackResponse:
    return MpesaCallbackResponse(**state.ingest_mpesa_callback(payload))


@router.post(
    "/print-jobs/transfers/{transfer_no}",
    response_model=PrintJobResponse,
    tags=["printing"],
)
def create_transfer_print_job(
    transfer_no: str,
    payload: DocumentPrintJobCreate,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return PrintJobResponse(**state.create_transfer_print_job(transfer_no, data))


@router.post(
    "/print-jobs/transfers/{transfer_no}/dispatch-bundle",
    response_model=TransferDispatchBundleResponse,
    tags=["printing", "transfers"],
)
def create_transfer_dispatch_bundle(
    transfer_no: str,
    payload: TransferDispatchBundleRequest,
    authorization: Optional[str] = Header(default=None),
) -> TransferDispatchBundleResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    result = state.create_transfer_dispatch_bundle(transfer_no, data)
    return TransferDispatchBundleResponse(
        transfer_no=result["transfer_no"],
        status=result.get("status", ""),
        delivery_batch_no=result.get("delivery_batch_no", ""),
        shipment_session_no=result.get("shipment_session_no", ""),
        transfer_print_job=PrintJobResponse(**result["transfer_print_job"]),
        label_print_jobs=[PrintJobResponse(**row) for row in result["label_print_jobs"]],
        total_label_copies=result["total_label_copies"],
        generated_bale_count=result.get("generated_bale_count", 0),
        store_dispatch_bales=[StoreDispatchBaleResponse(**row) for row in result.get("store_dispatch_bales", [])],
    )


@router.post(
    "/print-jobs/{job_id}/complete",
    response_model=PrintJobResponse | BaleLabelPrintJobResponse,
    tags=["printing"],
)
def mark_print_job_printed(
    job_id: int,
    payload: Optional[PrintStationCompleteRequest] = None,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse | BaleLabelPrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    if payload is not None:
        return BaleLabelPrintJobResponse(
            **state.complete_print_station_job(job_id, station_id=payload.station_id)
        )
    return PrintJobResponse(**state.mark_print_job_printed(job_id, printed_by=current_user["username"]))


@router.post(
    "/print-jobs/{job_id}/fail",
    response_model=PrintJobResponse | BaleLabelPrintJobResponse,
    tags=["printing"],
)
def mark_print_job_failed(
    job_id: int,
    payload: PrintJobFailureRequest | PrintStationFailRequest,
    authorization: Optional[str] = Header(default=None),
) -> PrintJobResponse | BaleLabelPrintJobResponse:
    current_user = _require_current_user(authorization=authorization)
    if isinstance(payload, PrintStationFailRequest):
        return BaleLabelPrintJobResponse(
            **state.fail_print_station_job(
                job_id,
                station_id=payload.station_id,
                error_message=payload.error_message,
            )
        )
    return PrintJobResponse(
        **state.mark_print_job_failed(
            job_id,
            failed_by=current_user["username"],
            note=payload.note,
        )
    )


@router.post("/sales", response_model=SaleResponse, tags=["sales"])
def create_sale(
    payload: SaleCreate,
    authorization: Optional[str] = Header(default=None),
) -> SaleResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["cashier_name"] = current_user["username"]
    return SaleResponse(**state.create_sale_transaction(data))


@router.get("/sales", response_model=list[SaleResponse], tags=["sales"])
def list_sales(
    authorization: Optional[str] = Header(default=None),
) -> list[SaleResponse]:
    _require_current_user(authorization=authorization)
    return [SaleResponse(**row) for row in state.list_sales_transactions()]


@router.get("/sales/void-requests", response_model=list[SaleVoidRequestResponse], tags=["sales"])
def list_sale_void_requests(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[SaleVoidRequestResponse]:
    _require_current_user(authorization=authorization)
    return [
        SaleVoidRequestResponse(**row)
        for row in state.list_sale_void_requests(store_code=store_code, status=status)
    ]


@router.post("/sales/{order_no}/void-request", response_model=SaleVoidRequestResponse, tags=["sales"])
def create_sale_void_request(
    order_no: str,
    payload: SaleVoidRequestCreate,
    authorization: Optional[str] = Header(default=None),
) -> SaleVoidRequestResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return SaleVoidRequestResponse(**state.create_sale_void_request(order_no, data))


@router.get("/sales/refund-requests", response_model=list[SaleRefundRequestResponse], tags=["sales"])
def list_sale_refund_requests(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[SaleRefundRequestResponse]:
    _require_current_user(authorization=authorization)
    return [
        SaleRefundRequestResponse(**row)
        for row in state.list_sale_refund_requests(store_code=store_code, status=status)
    ]


@router.post("/sales/{order_no}/refund-request", response_model=SaleRefundRequestResponse, tags=["sales"])
def create_sale_refund_request(
    order_no: str,
    payload: SaleRefundRequestCreate,
    authorization: Optional[str] = Header(default=None),
) -> SaleRefundRequestResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return SaleRefundRequestResponse(**state.create_sale_refund_request(order_no, data))


@router.post("/sales/refund-requests/{refund_no}/review", response_model=SaleRefundRequestResponse, tags=["sales"])
def review_sale_refund_request(
    refund_no: str,
    payload: SaleRefundReviewRequest,
    authorization: Optional[str] = Header(default=None),
) -> SaleRefundRequestResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["reviewed_by"] = current_user["username"]
    return SaleRefundRequestResponse(**state.review_sale_refund_request(refund_no, data))


@router.post("/sales/void-requests/{void_no}/review", response_model=SaleVoidRequestResponse, tags=["sales"])
def review_sale_void_request(
    void_no: str,
    payload: SaleVoidReviewRequest,
    authorization: Optional[str] = Header(default=None),
) -> SaleVoidRequestResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["reviewed_by"] = current_user["username"]
    return SaleVoidRequestResponse(**state.review_sale_void_request(void_no, data))


@router.post(
    "/sales/offline-sync",
    response_model=OfflineSaleSyncBatchResponse,
    tags=["sales", "offline"],
)
def sync_offline_sales(
    payload: OfflineSaleSyncBatchRequest,
    authorization: Optional[str] = Header(default=None),
) -> OfflineSaleSyncBatchResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["uploaded_by"] = current_user["username"]
    return OfflineSaleSyncBatchResponse(**state.sync_offline_sales_batch(data))


@router.get(
    "/sales/offline-sync-batches",
    response_model=list[OfflineSaleSyncBatchResponse],
    tags=["sales", "offline"],
)
def list_offline_sync_batches(
    device_id: Optional[str] = Query(default=None),
    store_code: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[OfflineSaleSyncBatchResponse]:
    _require_current_user(authorization=authorization)
    return [
        OfflineSaleSyncBatchResponse(**row)
        for row in state.list_offline_sync_batches(device_id=device_id, store_code=store_code)
    ]


@router.post("/pos/shifts/open", response_model=CashierShiftSummary, tags=["pos"])
def open_cashier_shift(
    payload: CashierShiftOpenRequest,
    authorization: Optional[str] = Header(default=None),
) -> CashierShiftSummary:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["opened_by"] = current_user["username"]
    return CashierShiftSummary(**state.open_cashier_shift(data))


@router.get("/pos/shifts", response_model=list[CashierShiftSummary], tags=["pos"])
def list_cashier_shifts(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[CashierShiftSummary]:
    _require_current_user(authorization=authorization)
    return [CashierShiftSummary(**row) for row in state.list_cashier_shifts(store_code=store_code, status=status)]


@router.post("/pos/shifts/{shift_no}/close", response_model=CashierShiftSummary, tags=["pos"])
def close_cashier_shift(
    shift_no: str,
    payload: CashierShiftCloseRequest,
    authorization: Optional[str] = Header(default=None),
) -> CashierShiftSummary:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["closed_by"] = current_user["username"]
    return CashierShiftSummary(**state.close_cashier_shift(shift_no, data))


@router.post("/pos/shifts/{shift_no}/handover-request", response_model=CashierHandoverLogResponse, tags=["pos"])
def request_cashier_handover(
    shift_no: str,
    payload: CashierHandoverRequest,
    authorization: Optional[str] = Header(default=None),
) -> CashierHandoverLogResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["requested_by"] = current_user["username"]
    return CashierHandoverLogResponse(**state.request_cashier_handover(shift_no, data))


@router.post("/pos/handovers/{handover_no}/review", response_model=CashierHandoverLogResponse, tags=["pos"])
def review_cashier_handover(
    handover_no: str,
    payload: CashierHandoverReviewRequest,
    authorization: Optional[str] = Header(default=None),
) -> CashierHandoverLogResponse:
    current_user = _require_current_user(authorization=authorization)
    data = payload.model_dump()
    data["reviewed_by"] = current_user["username"]
    return CashierHandoverLogResponse(**state.review_cashier_handover(handover_no, data))


@router.get("/pos/handovers", response_model=list[CashierHandoverLogResponse], tags=["pos"])
def list_cashier_handovers(
    store_code: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    authorization: Optional[str] = Header(default=None),
) -> list[CashierHandoverLogResponse]:
    _require_current_user(authorization=authorization)
    return [
        CashierHandoverLogResponse(**row)
        for row in state.list_cashier_handover_logs(store_code=store_code, status=status)
    ]


@router.get("/pos/shifts/{shift_no}/t-report", response_model=PosReportResponse, tags=["pos"])
def get_t_report(
    shift_no: str,
    authorization: Optional[str] = Header(default=None),
) -> PosReportResponse:
    _require_current_user(authorization=authorization)
    return PosReportResponse(**state.get_cashier_shift_report(shift_no, "t_report"))


@router.get("/pos/shifts/{shift_no}/z-report", response_model=PosReportResponse, tags=["pos"])
def get_z_report(
    shift_no: str,
    authorization: Optional[str] = Header(default=None),
) -> PosReportResponse:
    _require_current_user(authorization=authorization)
    return PosReportResponse(**state.get_cashier_shift_report(shift_no, "z_report"))
